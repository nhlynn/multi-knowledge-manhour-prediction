"""Export route blueprint for MHES.

Handles data export to Excel format matching the simple_resource template.

Generated workbooks are staged locally only long enough to be uploaded to
Google Cloud Storage (see services/gcs_service.py) — the local file is
deleted right after upload, and the bucket is the only persistent copy
from then on. Downloads are served via short-lived signed URLs instead of
streaming the file from local disk.

Export history rows created before this migration still have a local
absolute path in ``file_path`` — ``is_local_path`` (services.gcs_service)
tells those apart from the GCS object paths new rows use, so old records
keep working unchanged.
"""

import logging
import io
import os
import re
from datetime import datetime

from flask import (
    Blueprint,
    current_app,
    flash,
    jsonify,
    redirect,
    render_template,
    request,
    send_file,
    session,
    url_for,
)
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from services.export_history_service import ExportHistoryService
from services.gcs_service import (
    GCSConflictError,
    GCSError,
    blob_exists,
    download_excel_bytes,
    generate_signed_download_url,
    is_local_path,
    upload_excel_to_gcs,
)
from services.remark_html import build_single_cell_data, remark_html_to_lines, sanitize_remark_html
from utils.permissions import require_login

logger = logging.getLogger(__name__)

export_bp = Blueprint("export", __name__)
# Any logged-in role (Member and above) can export results.
export_bp.before_request(require_login)


def _team_id_filter() -> int | None:
    """Return the ``team_id`` to scope Export History reads to (Phase 6).

    None means "no filter" (see every team's exports) — Admin only.
    Every other role only ever sees their own team's exports.
    """
    if session.get("role") == "Admin":
        return None
    return session.get("team_id")


def _team_export_template() -> dict:
    """Return the current session's team's export template (Phase 8).

    Falls back to ``DEFAULT_EXPORT_TEMPLATE`` — which reproduces the
    exact pre-Phase-8 column layout — for any team with no configured
    template, so existing exports are completely unaffected.
    """
    from repositories.team_export_template_repository import TeamExportTemplateRepository

    repo = TeamExportTemplateRepository(current_app.config["MHES_DB_PATH"])
    config = repo.get_by_team_id(session["team_id"])
    return config["template_config"] if config else DEFAULT_EXPORT_TEMPLATE


@export_bp.route("/excel", methods=["POST"])
def export_excel():
    """Export preview data to an Excel file, stored in Google Cloud Storage.

    Expects JSON body with ``projectName`` and ``categories``. Builds the
    workbook to a temporary local file, uploads it to GCS, deletes the
    temp file, records the export in the history database, and returns
    the file as a download (served from the in-memory bytes already read
    for the upload, so no second disk read is needed).
    """
    data = request.get_json(silent=True) or {}
    project_name = (data.get("projectName") or "").strip()
    created_by = (data.get("createdBy") or "").strip()
    project_remark = sanitize_remark_html(data.get("projectRemark") or "")
    categories = data.get("categories", [])

    if not project_name:
        return jsonify({"error": "Project name is required."}), 400

    if not created_by:
        return jsonify({"error": "Created By is required."}), 400

    if not categories:
        return jsonify({"error": "No data to export."}), 400

    safe_name = re.sub(r'[\\/*?:"<>|]', "_", project_name)
    temp_dir = _export_folder()
    os.makedirs(temp_dir, exist_ok=True)
    build_path = os.path.join(temp_dir, _timestamped_export_filename(safe_name))

    try:
        _build_workbook(
            build_path, project_name, created_by, project_remark, categories,
            template_config=_team_export_template(),
        )
        with open(build_path, "rb") as f:
            file_bytes = f.read()
    except Exception as e:
        logger.error(f"Export failed: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        try:
            os.remove(build_path)
        except OSError:
            logger.warning("Could not remove temporary export file: %s", build_path)

    try:
        filename, object_path = _upload_export_with_retry(temp_dir, safe_name, file_bytes)
    except GCSError as e:
        logger.exception("Failed to upload export file to GCS for project=%r.", project_name)
        return jsonify({"error": str(e)}), 502

    logger.info(
        "Export succeeded: file=%s project_name=%r gcs_path=%s",
        filename, project_name, object_path,
    )
    _record_export_history(
        object_path, len(file_bytes), filename, project_name, created_by, categories,
        team_id=session["team_id"], created_by_user_id=session.get("user_id"),
    )

    return send_file(
        io.BytesIO(file_bytes),
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def _timestamped_export_filename(safe_name: str) -> str:
    """Build an export filename with a millisecond-precision timestamp
    suffix, so repeat exports of the same project never collide/overwrite
    each other in GCS (object paths are keyed by file_name — see
    services/gcs_service.py). Colons/periods aren't valid in Windows
    filenames, so dd-mm-yyyy_HH-mm-ss-SSS is used instead of
    dd-mm-yyyy HH:mm:ss.SSS.
    """
    timestamp = datetime.now().strftime("%d-%m-%Y_%H-%M-%S-%f")[:-3]
    return f"{safe_name}_manhour_{timestamp}.xlsx"


def _upload_export_with_retry(
    temp_dir: str, safe_name: str, file_bytes: bytes, max_attempts: int = 3,
) -> tuple[str, str]:
    """Upload the given Excel bytes to GCS under a fresh timestamped
    filename, retrying with a new timestamp if GCS rejects the write
    because an object already exists at that exact path — an extremely
    unlikely millisecond-timestamp collision (e.g. two exports of the same
    project landing in the same millisecond), rejected by GCS itself via
    ``upload_excel_to_gcs``'s ``if_generation_match=0`` precondition rather
    than silently overwritten. Transparent to the caller: on success this
    just looks like one upload with a slightly later timestamp.

    Returns:
        ``(filename, object_path)`` for the attempt that succeeded.

    Raises:
        GCSError: The error from the final attempt, if every attempt failed.
    """
    last_error: GCSError | None = None
    for attempt in range(1, max_attempts + 1):
        filename = _timestamped_export_filename(safe_name)
        temp_path = os.path.join(temp_dir, filename)
        try:
            with open(temp_path, "wb") as f:
                f.write(file_bytes)
            object_path = upload_excel_to_gcs(temp_path, filename)
            return filename, object_path
        except GCSConflictError as e:
            last_error = e
            logger.warning(
                "Export filename collided in GCS (attempt %d/%d): %s. Retrying with a new timestamp.",
                attempt, max_attempts, filename,
            )
        except GCSError as e:
            last_error = e
            break
        finally:
            try:
                os.remove(temp_path)
            except OSError:
                logger.warning("Could not remove temporary export file: %s", temp_path)
    raise last_error


def _record_export_history(
    object_path: str, file_size: int, filename: str, project_name: str, created_by: str, categories: list,
    *, team_id: int, created_by_user_id: int | None,
) -> None:
    """Save export metadata to the Export History database.

    Best-effort: the Excel file was already uploaded to GCS successfully
    by this point, so a failure here is only logged — it must never
    remove the uploaded file or fail the export response.

    Args:
        object_path: The GCS object path returned by ``upload_excel_to_gcs``
            (e.g. "mhes/bcmm/1001/estimate_001.xlsx"), stored as
            ``file_path``.
        file_size: Size, in bytes, of the generated Excel file.
        team_id: The exporting user's team (Phase 6) — ``session["team_id"]``.
        created_by_user_id: The exporting user's id (Phase 6) —
            ``session.get("user_id")``.
    """
    try:
        total_tasks = sum(len(cat.get("tasks") or []) for cat in categories)
        total_hours = sum(
            (task.get("total_hours") or 0)
            for cat in categories
            for task in (cat.get("tasks") or [])
        )
        _export_history_service().insert_history(
            project_name=project_name,
            created_by=created_by,
            created_by_user_id=created_by_user_id,
            team_id=team_id,
            export_date=datetime.now().isoformat(),
            file_name=filename,
            file_url=url_for("export.download_export", filename=filename),
            file_path=object_path,
            file_size=file_size,
            total_tasks=total_tasks,
            total_hours=total_hours,
        )
    except Exception:
        logger.exception(
            "Failed to save export history for file=%s; the Excel file was still uploaded to GCS.",
            filename,
        )


def _export_folder() -> str:
    """Local scratch directory used only to stage a workbook before it's
    uploaded to GCS — files here are temporary and deleted right after
    upload (see export_excel). Not where exports are persisted anymore.
    """
    return os.path.join(current_app.root_path, "exports")


def _export_history_service() -> ExportHistoryService:
    return ExportHistoryService(db_path=current_app.config["MHES_DB_PATH"])


EXPORTS_PER_PAGE = 10


@export_bp.route("/files", methods=["GET"])
def list_exports() -> str:
    """Render the Export History page from SQLite metadata (no folder scan).

    Supports server-side pagination (``page``) combined with the From
    Date / To Date / Project Name filters, so only one page of records is
    ever loaded from the database per request.
    """
    # On a fresh visit (no query string at all) default both date fields to
    # today, so the page opens already scoped to "today's exports" instead
    # of loading the entire history. Once the user has interacted with the
    # filter form — including clearing the dates via Reset, which passes
    # explicit empty values — the field is present in the query string
    # (even if empty), so the default is not re-applied and their choice
    # (including "no date filter") is respected.
    today_str = datetime.now().strftime("%Y-%m-%d")
    from_date = (request.args.get("from_date", today_str) or "").strip()
    to_date = (request.args.get("to_date", today_str) or "").strip()
    project_name = (request.args.get("project_name") or "").strip()
    try:
        page = int(request.args.get("page", 1))
    except ValueError:
        page = 1
    page = max(page, 1)

    service = _export_history_service()
    team_filter = _team_id_filter()
    try:
        history, total = service.get_history_page(
            page=page,
            per_page=EXPORTS_PER_PAGE,
            team_id=team_filter,
            from_date=from_date or None,
            to_date=to_date or None,
            project_name=project_name or None,
        )
        total_pages = max((total + EXPORTS_PER_PAGE - 1) // EXPORTS_PER_PAGE, 1)
        if page > total_pages:
            # Requested page is past the last one (e.g. stale bookmark
            # after records were deleted) — re-fetch the actual last page
            # instead of showing an empty table under a wrong page number.
            page = total_pages
            history, total = service.get_history_page(
                page=page,
                per_page=EXPORTS_PER_PAGE,
                team_id=team_filter,
                from_date=from_date or None,
                to_date=to_date or None,
                project_name=project_name or None,
            )
    except Exception:
        logger.exception("Failed to load export history from database.")
        history, total, total_pages = [], 0, 1

    # Admin sees exports across every team — enrich each row with the
    # owning team's name so that's visible in the list (team_filter is
    # None only for Admin; see _team_id_filter).
    team_names_by_id = {}
    if team_filter is None and history:
        from repositories.team_repository import TeamRepository

        team_names_by_id = {
            t["id"]: t["name"]
            for t in TeamRepository(current_app.config["MHES_DB_PATH"]).list_all()
        }

    export_folder = _export_folder()
    for record in history:
        if team_filter is None:
            record["team_name"] = team_names_by_id.get(record.get("team_id"), "Unknown")
        file_path = record.get("file_path")
        if file_path and is_local_path(file_path):
            # Pre-migration row — still a real local export, check disk directly.
            record["file_exists"] = os.path.isfile(file_path)
        elif file_path:
            # Post-migration row — file_path is a GCS object path.
            record["file_exists"] = blob_exists(file_path)
        else:
            # Oldest rows, from before the file_path column existed at
            # all — fall back to reconstructing the (local) path.
            local_path = os.path.join(export_folder, record["file_name"])
            record["file_exists"] = os.path.isfile(local_path)
        record["size_kb"] = round(record["file_size"] / 1024, 1) if record.get("file_size") else 0
        if not record["file_exists"]:
            logger.warning(
                "Export history file missing: %s (history id=%s)",
                record["file_name"], record["id"],
            )

    range_start = (page - 1) * EXPORTS_PER_PAGE + 1 if total else 0
    range_end = min(page * EXPORTS_PER_PAGE, total)

    filter_args = {}
    if from_date:
        filter_args["from_date"] = from_date
    if to_date:
        filter_args["to_date"] = to_date
    if project_name:
        filter_args["project_name"] = project_name

    return render_template(
        "exported_files.html",
        files=history,
        page=page,
        total_pages=total_pages,
        total=total,
        range_start=range_start,
        range_end=range_end,
        from_date=from_date,
        to_date=to_date,
        project_name=project_name,
        filter_args=filter_args,
        has_filters=bool(filter_args),
    )


def _is_safe_export_filename(filename: str) -> bool:
    """Reject path traversal / directory separators without mangling the
    filename itself (unlike ``werkzeug.secure_filename``, which replaces
    spaces with underscores and would no longer match real export
    filenames — these are generated by this app, e.g. "Project 1_manhour.xlsx",
    not arbitrary user-uploaded names, so we only need to guard against
    escaping the exports folder, not normalize the name).
    """
    return (
        bool(filename)
        and filename.lower().endswith(".xlsx")
        and os.path.basename(filename) == filename
        and ".." not in filename
    )


@export_bp.route("/files/<filename>", methods=["GET"])
def download_export(filename: str):
    """Download a previously exported Excel file.

    Looks up where the file actually lives via its export_history row: a
    GCS object path for exports created after the storage migration (in
    which case the browser is redirected to a short-lived signed URL, so
    the file is streamed directly from GCS rather than through this
    server), or a local absolute path for older, pre-migration exports
    (served directly from disk, as before).
    """
    if not _is_safe_export_filename(filename):
        logger.warning("Rejected unsafe export filename for download: %r", filename)
        flash(f"File not found: {filename}", "warning")
        return redirect(url_for("export.list_exports"))

    record = _export_history_service().get_history_by_file_name(filename, team_id=_team_id_filter())
    if record is None:
        # Either the file truly doesn't exist, or it exists but belongs
        # to another team (Phase 6) — treated identically as "not found".
        # No local-path-reconstruction fallback here: guessing a path
        # for an unscoped record could otherwise leak a legacy local
        # export across teams.
        logger.warning("Export history record not found or not owned by caller's team: %s", filename)
        flash(f"File not found: {filename}", "warning")
        return redirect(url_for("export.list_exports"))

    # Rows created before the file_path column existed at all have a
    # legitimate record but no file_path — reconstruct the (local) path
    # the same way list_exports does for those.
    file_path = record.get("file_path") or os.path.join(_export_folder(), filename)

    if is_local_path(file_path):
        if not os.path.isfile(file_path):
            logger.warning("Export file missing on disk for download: %s", filename)
            flash(f"File not found: {filename}", "warning")
            return redirect(url_for("export.list_exports"))
        try:
            return send_file(file_path, as_attachment=True, download_name=filename)
        except Exception:
            logger.exception("Failed to send export file for download: %s", filename)
            flash(f"Could not download '{filename}'.", "danger")
            return redirect(url_for("export.list_exports"))

    try:
        signed_url = generate_signed_download_url(file_path, download_name=filename)
    except GCSError:
        logger.exception("Failed to generate signed download URL for: %s", filename)
        flash(f"Could not download '{filename}'.", "danger")
        return redirect(url_for("export.list_exports"))

    return redirect(signed_url)


@export_bp.route("/files/<filename>/view", methods=["GET"])
def view_export(filename: str) -> str:
    """Render a read-only in-browser detail view of an exported file.

    Reads the workbook from GCS (post-migration exports) or local disk
    (pre-migration exports), based on the export_history row's file_path —
    same local-vs-GCS distinction as ``download_export``.
    """
    if not _is_safe_export_filename(filename):
        flash(f"File not found: {filename}", "warning")
        return redirect(url_for("export.list_exports"))

    record = _export_history_service().get_history_by_file_name(filename, team_id=_team_id_filter())
    if record is None:
        # Either the file truly doesn't exist, or it belongs to another
        # team (Phase 6) — treated identically as "not found".
        flash(f"File not found: {filename}", "warning")
        return redirect(url_for("export.list_exports"))

    file_path = record.get("file_path") or os.path.join(_export_folder(), filename)

    try:
        if is_local_path(file_path):
            if not os.path.isfile(file_path):
                flash(f"File not found: {filename}", "warning")
                return redirect(url_for("export.list_exports"))
            detail = _read_export_detail(file_path)
        else:
            file_bytes = download_excel_bytes(file_path)
            detail = _read_export_detail(io.BytesIO(file_bytes))
    except GCSError as e:
        logger.exception(f"Failed to download export file from GCS for '{filename}'.")
        flash(f"Could not open '{filename}' for viewing: {e}", "danger")
        return redirect(url_for("export.list_exports"))
    except Exception as e:
        logger.exception(f"Failed to read export detail for '{filename}'.")
        flash(f"Could not open '{filename}' for viewing: {e}", "danger")
        return redirect(url_for("export.list_exports"))

    return render_template("export_detail.html", filename=filename, **detail)


def _read_export_detail(filepath) -> dict:
    """Read an exported workbook back into a display-friendly structure.

    ``filepath`` may be a path string (local disk) or a file-like object
    such as ``io.BytesIO`` (downloaded from GCS) — ``openpyxl.load_workbook``
    accepts either.

    Mirrors the exact row layout ``_build_workbook`` writes (title,
    Created By/Date, headers, category/task rows, Total, Remark), so this
    only works for files this app generated itself — which is the only
    kind that ever lands in the exports folder.
    """
    from openpyxl import load_workbook

    wb = load_workbook(filepath, rich_text=True)
    try:
        ws = wb.active

        title = (ws["A1"].value or "").strip()
        project_name = re.sub(r"\s*Manhour\s*$", "", title).strip() or title
        created_by = ws["E3"].value or ""
        date_value = ws["E4"].value
        date_str = date_value.strftime("%Y-%m-%d") if hasattr(date_value, "strftime") else (date_value or "")

        categories = []
        total_row = None
        r = 6
        while True:
            cat_value = ws.cell(row=r, column=1).value
            task_value = ws.cell(row=r, column=2).value

            if isinstance(cat_value, str) and cat_value.strip() == "Total":
                total_row = r
                break
            if cat_value is None and task_value is None:
                # No more data rows (shouldn't normally happen before a
                # Total row, but avoids an infinite loop on a malformed file).
                break

            if cat_value:
                categories.append({"name": cat_value, "rows": []})
            if not categories:
                categories.append({"name": "", "rows": []})

            estimate = ws.cell(row=r, column=3).value
            working_day = round(estimate / 8, 2) if isinstance(estimate, (int, float)) else ""
            categories[-1]["rows"].append({
                "task": task_value or "",
                "estimate": estimate,
                "working_day": working_day,
                "remarks": ws.cell(row=r, column=5).value or "",
            })
            r += 1

        grand_total = ws.cell(row=total_row, column=3).value if total_row else 0
        grand_working_day = round(grand_total / 8, 2) if isinstance(grand_total, (int, float)) else ""

        remark_row = (total_row + 3) if total_row else None
        remark_cell = ws.cell(row=remark_row, column=1) if remark_row else None
        remark_html = _rich_text_to_html(remark_cell.value if remark_cell else None)
        remark_bg = _fill_to_hex(remark_cell.fill) if remark_cell else None
        remark_hyperlink = remark_cell.hyperlink.target if (remark_cell and remark_cell.hyperlink) else None

        return {
            "project_name": project_name,
            "created_by": created_by,
            "date_str": date_str,
            "categories": categories,
            "grand_total": grand_total,
            "grand_working_day": grand_working_day,
            "remark_html": remark_html,
            "remark_bg": remark_bg,
            "remark_hyperlink": remark_hyperlink,
        }
    finally:
        wb.close()


def _fill_to_hex(fill) -> str | None:
    """Convert an openpyxl cell fill to a CSS hex color, or None if unfilled."""
    if not fill or fill.fill_type != "solid" or not fill.fgColor:
        return None
    rgb = fill.fgColor.rgb
    if isinstance(rgb, str) and len(rgb) == 8:
        return "#" + rgb[2:]
    return None


def _rich_text_to_html(value) -> str:
    """Convert a cell's rich-text (or plain string) value into safe HTML.

    Only for read-only display of our own generated files — reconstructs
    bold/italic/underline/font color and line breaks from the openpyxl
    rich-text runs. This is the display-side mirror of
    ``services/remark_html.py``'s HTML-to-rich-text conversion.
    """
    from html import escape

    from openpyxl.cell.rich_text import CellRichText

    if value in (None, "") or (isinstance(value, str) and value.strip() == "No remark added."):
        return '<span class="text-muted fst-italic">No remark added.</span>'

    if isinstance(value, str):
        return escape(value).replace("\n", "<br>")

    runs = value if isinstance(value, CellRichText) else [value]
    html_parts = []
    for part in runs:
        text = part.text if hasattr(part, "text") else str(part)
        escaped = escape(text).replace("\n", "<br>")
        font = getattr(part, "font", None)
        if font is None:
            html_parts.append(escaped)
            continue

        color_hex = None
        if font.color and isinstance(font.color.rgb, str) and len(font.color.rgb) == 8:
            color_hex = "#" + font.color.rgb[2:]

        open_tags, close_tags = "", ""
        if font.b:
            open_tags += "<strong>"
            close_tags = "</strong>" + close_tags
        if font.i:
            open_tags += "<em>"
            close_tags = "</em>" + close_tags
        if font.u:
            open_tags += "<u>"
            close_tags = "</u>" + close_tags

        style_attr = f' style="color: {color_hex};"' if color_hex else ""
        html_parts.append(f"<span{style_attr}>{open_tags}{escaped}{close_tags}</span>")

    return "".join(html_parts)


DEFAULT_EXPORT_TEMPLATE = {
    "sheet_title": "Manhour",
    "columns": [
        {"key": "category", "label": "Category", "width": 25},
        {"key": "task", "label": "Task List", "width": 45},
        {"key": "estimate_hours", "label": "Estimate (Hours)", "width": 22},
        {"key": "working_day", "label": "Working Day", "width": 15},
        {"key": "remarks", "label": "Remarks", "width": 35},
    ],
}
"""The pre-Phase-8 column layout, reproduced exactly as data instead of
hardcoded Excel column letters — this is what every team without a
configured ``team_export_templates`` row gets, so existing exports are
byte-for-byte unaffected by Phase 8 (see docs/ARCHITECTURE.md §5h).

Recognized ``columns[].key`` values (each renders one data-table column;
unknown keys render as an empty column and are logged):
    category        -- category name, merged across that category's task rows
    task            -- numbered task name ("1. <task>")
    estimate_hours  -- the task's total hours
    working_day     -- formula: whichever "estimate_hours" column's value / 8
                       (blank if the template has no "estimate_hours" column)
    remarks         -- free-text remarks (task.get("remarks", ""))
"""


def _build_workbook(
    filepath: str,
    project_name: str,
    created_by: str,
    project_remark: str,
    categories: list,
    template_config: dict | None = None,
) -> None:
    """Build an Excel workbook using a team's configured column template.

    Args:
        template_config: Optional per-team template (Phase 8 — see
            ``DEFAULT_EXPORT_TEMPLATE`` above and
            ``repositories/team_export_template_repository.py``). None
            uses ``DEFAULT_EXPORT_TEMPLATE``, reproducing the exact
            pre-Phase-8 layout.

    The title/Created-By/Date metadata rows, per-category row merging,
    totals row, and the rich-text Remark section are shared structure —
    identical for every team regardless of template; only the data
    table's columns (which ones appear, their order, label, and width)
    are configurable.
    """
    template_config = template_config or DEFAULT_EXPORT_TEMPLATE
    columns = template_config.get("columns") or DEFAULT_EXPORT_TEMPLATE["columns"]
    sheet_title = template_config.get("sheet_title") or DEFAULT_EXPORT_TEMPLATE["sheet_title"]
    num_cols = len(columns)
    col_index_by_key = {col["key"]: i for i, col in enumerate(columns, 1)}
    category_col = col_index_by_key.get("category")
    estimate_col = col_index_by_key.get("estimate_hours")
    working_day_col = col_index_by_key.get("working_day")

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title

    # --- Styles ---
    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    cat_font = Font(bold=True)
    total_font = Font(bold=True)
    total_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    center_align = Alignment(horizontal="center", vertical="center")
    wrap_align = Alignment(vertical="center", wrap_text=True)

    # Column widths
    for i, col in enumerate(columns, 1):
        if col.get("width"):
            ws.column_dimensions[get_column_letter(i)].width = col["width"]

    # --- Row 1-2: Title (merged across every configured column) ---
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=num_cols)
    title_cell = ws.cell(row=1, column=1, value=f"{project_name} {sheet_title}")
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # --- Rows 3-4: Created By / Date (label in the second-to-last
    # column, value in the last — generalizes the original D3/E3, D4/E4
    # for whatever column count this template has) ---
    label_col = max(num_cols - 1, 1)
    value_col = num_cols
    ws.cell(row=3, column=label_col, value="Created By")
    created_by_cell = ws.cell(row=3, column=value_col, value=created_by)
    created_by_cell.alignment = Alignment(horizontal="left", vertical="center")

    ws.cell(row=4, column=label_col, value="Date")
    date_cell = ws.cell(row=4, column=value_col, value=datetime.now())
    date_cell.number_format = r"yyyy\-mm\-dd"
    date_cell.alignment = Alignment(horizontal="left", vertical="center")

    # --- Row 5: Headers ---
    for i, col in enumerate(columns, 1):
        cell = ws.cell(row=5, column=i, value=col.get("label") or col["key"])
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # --- Data rows ---
    row = 6
    grand_total = 0

    for cat in categories:
        cat_start_row = row
        cat_name = cat.get("category", "")

        # Each task as a numbered row (no activity detail flattening)
        task_num = 1
        cat_total_hours = 0

        for task in cat.get("tasks", []):
            task_name = task.get("task", "")
            total_hours = task.get("total_hours", 0)

            for i, col in enumerate(columns, 1):
                key = col["key"]
                if key == "category":
                    continue  # written once per category block, after this loop
                elif key == "task":
                    cell = ws.cell(row=row, column=i, value=f"{task_num}. {task_name}")
                    cell.alignment = wrap_align
                elif key == "estimate_hours":
                    cell = ws.cell(row=row, column=i, value=total_hours)
                    cell.alignment = center_align
                elif key == "working_day":
                    value = f"={get_column_letter(estimate_col)}{row}/8" if estimate_col else None
                    cell = ws.cell(row=row, column=i, value=value)
                    cell.alignment = center_align
                elif key == "remarks":
                    cell = ws.cell(row=row, column=i, value=task.get("remarks", ""))
                    cell.alignment = wrap_align
                else:
                    logger.warning("Unknown export template column key %r; left blank.", key)
                    cell = ws.cell(row=row, column=i)
                cell.border = thin_border

            cat_total_hours += total_hours
            task_num += 1
            row += 1

        cat_end_row = row - 1
        if cat_end_row < cat_start_row:
            continue

        # Category column (merged), if this template has one
        if category_col:
            cat_row_count = cat_end_row - cat_start_row + 1
            if cat_row_count > 1:
                ws.merge_cells(
                    start_row=cat_start_row, start_column=category_col,
                    end_row=cat_end_row, end_column=category_col,
                )
            cat_cell = ws.cell(row=cat_start_row, column=category_col, value=cat_name)
            cat_cell.font = cat_font
            cat_cell.alignment = Alignment(vertical="center")
            cat_cell.border = thin_border
            for r in range(cat_start_row, cat_end_row + 1):
                ws.cell(row=r, column=category_col).border = thin_border

        grand_total += cat_total_hours

    # --- Total row ---
    total_row = row
    ws.cell(row=total_row, column=1, value="Total").font = total_font
    for col_idx in range(1, num_cols + 1):
        cell = ws.cell(row=total_row, column=col_idx)
        cell.fill = total_fill
        cell.border = thin_border
        cell.font = total_font

    if estimate_col:
        ws.cell(row=total_row, column=estimate_col, value=grand_total).alignment = center_align
    if working_day_col and estimate_col:
        ws.cell(
            row=total_row, column=working_day_col,
            value=f"={get_column_letter(estimate_col)}{total_row}/8",
        ).alignment = center_align

    # --- Remark section ---
    # The whole remark lives in ONE merged cell/row (auto-sized height)
    # rather than one row per line. Bold/italic/underline/font color and
    # bullet/numbered-list markers are preserved exactly, per character.
    # A single cell can only carry one fill color and one hyperlink for
    # its entire content though, so if the remark uses more than one
    # highlight color or more than one link, only the first of each
    # applies to the whole cell (see services/remark_html.py for why).
    # This section is shared structure, identical for every team's
    # template — only the merge span (num_cols) varies.
    remark_header_row = total_row + 2
    ws.merge_cells(
        start_row=remark_header_row, start_column=1,
        end_row=remark_header_row, end_column=num_cols,
    )
    remark_header_cell = ws.cell(row=remark_header_row, column=1, value="Remark:")
    remark_header_cell.font = cat_font
    # A border set only on the merged range's top-left cell only draws
    # that one cell's edges — Excel needs every underlying cell in the
    # merge to carry the border, or the other sides (right/bottom here)
    # are left open. Same reasoning applies to the remark content row below.
    for col_idx in range(1, num_cols + 1):
        ws.cell(row=remark_header_row, column=col_idx).border = thin_border

    remark_row = remark_header_row + 1
    remark_lines = remark_html_to_lines(project_remark)
    cell_data = build_single_cell_data(remark_lines)

    line_count = max(len(remark_lines), 1)
    row_height = min(409, max(20, line_count * 15 + 5))
    ws.merge_cells(start_row=remark_row, start_column=1, end_row=remark_row, end_column=num_cols)
    ws.row_dimensions[remark_row].height = row_height

    for col_idx in range(1, num_cols + 1):
        ws.cell(row=remark_row, column=col_idx).border = thin_border

    remark_cell = ws.cell(row=remark_row, column=1)
    remark_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    if cell_data is None:
        remark_cell.value = "No remark added."
        remark_cell.font = Font(italic=True, color="94A3B8")
    else:
        remark_cell.value = cell_data["value"]
        if cell_data["fill"]:
            remark_cell.fill = PatternFill(
                start_color=cell_data["fill"], end_color=cell_data["fill"], fill_type="solid"
            )
        if cell_data["hyperlink"]:
            remark_cell.hyperlink = cell_data["hyperlink"]

    wb.save(filepath)
