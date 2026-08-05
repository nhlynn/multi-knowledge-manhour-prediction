"""One-off script that generated ``import/bamawl/bamawl_import_template.xlsx``
from ``simple_resource/bamawl_import_export_format_filled.xlsx``.

Not part of the application's runtime import/export code paths (see
``services/team_template_validator.py`` / ``services/bamawl_export_builder.py``
for those) -- this is a standalone, one-time sanitization script, kept
alongside its output for reproducibility/traceability (e.g. if the real
source template's structure ever changes and this sample needs
regenerating).

Replaces every real-project/customer-identifying value in the source
workbook with generic sample text or round sample numbers, while
leaving the workbook's structure, worksheet names, formulas, data
validation, merged cells, formatting, row heights, and column widths
completely untouched -- only specific cells' *values* are replaced.

Known, unavoidable characteristic (an openpyxl limitation, not
something this script introduces): any workbook re-saved via openpyxl
loses the *cached* display value of every formula cell workbook-wide,
regardless of which cells were actually edited. The formulas
themselves are preserved verbatim, and opening this file in real
Microsoft Excel recalculates and displays correct sample values
immediately (Excel's default calculation mode is Automatic) -- this
only matters for a tool that reads cached values without recalculating
(e.g. pandas) before the file has ever been opened/saved in Excel.
"""

import openpyxl

SOURCE = "simple_resource/bamawl_import_export_format_filled.xlsx"
DEST = "import/bamawl/bamawl_import_template.xlsx"

wb = openpyxl.load_workbook(SOURCE)

# --- ReqDefinition: generic sample text for every section ---
req = wb["ReqDefinition"]
req["B1"] = "Sample HR & Attendance Management System"
req["C3"] = "Sample purpose description: digitize employee attendance and leave management. For demonstration purposes only."
req["C4"] = "Sample scope: employee master data management, attendance tracking, leave request workflow, and reporting. For demonstration purposes only."
req["C5"] = "Sample document specifications: design document, test specification, user manual. For demonstration purposes only."
req["C6"] = "Sample functional specifications: role-based login, master data registers, approval workflow, notifications. For demonstration purposes only."
req["C7"] = "Sample server environment: on-premise server with standard specifications. For demonstration purposes only."
req["C12"] = "Sample out-of-scope items: payroll processing, third-party integrations, mobile application. For demonstration purposes only."

# --- FunctionList: realistic-sounding but clearly generic sample
# function names, same row count. Not the real project's actual
# feature list -- distinct, plausible names for a demonstration. ---
SAMPLE_FUNCTION_NAMES = [
    "Sample User Login", "Sample User Registration", "Sample Profile Management",
    "Sample Role & Permission Setup", "Sample Department Master", "Sample Item Master",
    "Sample Data Import", "Sample Data Export", "Sample Record Search",
    "Sample Record Approval Flow", "Sample Record Rejection Flow", "Sample Status Dashboard",
    "Sample Calendar Master", "Sample Notification Setting", "Sample Activity Log",
    "Sample Report Generation (CSV/PDF)", "Sample Summary Report", "Sample Account Settings",
    "Sample Password Reset", "Sample Announcement Board", "Sample Health Monitoring",
    "Sample Admin Dashboard", "Sample Audit Trail", "Sample System Configuration",
]
function_list = wb["FunctionList"]
for row, name in enumerate(SAMPLE_FUNCTION_NAMES, start=2):
    function_list.cell(row=row, column=3, value=name)

# --- TotalManhour: generic labels for the device-specific rows only;
# every other label, all numbers, and every formula are left as-is. ---
total_manhour = wb["TotalManhour"]
total_manhour["B10"] = "Sample Device Test 1"
total_manhour["B11"] = "Sample Device Test 2"
total_manhour["B12"] = "Sample Device Test 3"

# --- Infra Manhour: remove the real project title and real customer
# name; the generic vendor-requirements boilerplate (VM/OS/firewall/
# backup bullet points) isn't customer-identifying and is left as-is. ---
infra = wb["Infra Manhour"]
infra["A1"] = "Man Hour for Sample HR & Attendance System"
infra["B4"] = "Sample Customer Co., Ltd."

# --- ALL_Detail: the knowledge source -- only ID/Function/Status/
# Development-hours (columns A-D) are literal values; replace those
# with sample data. Every formula column (E onward, including the
# capacity-planning block below the task rows and the Required Skill
# section's specific tech-stack mention) keeps its exact formula text,
# per "keep all formulas" -- only the two Required Skill VALUE cells
# (not formulas) are genericized. ---
all_detail = wb["ALL_Detail"]
# Same realistic-but-generic names as the first 9 FunctionList entries,
# so a reader studying both sheets sees a consistent, coherent example
# rather than two unrelated naming schemes.
sample_rows = list(zip(SAMPLE_FUNCTION_NAMES[:9], [8, 16, 24, 12, 20, 10, 18, 14, 16]))
for i, (name, hours) in enumerate(sample_rows, start=1):
    row = 4 + i  # rows 5-13
    all_detail.cell(row=row, column=1, value=i)  # ID
    all_detail.cell(row=row, column=2, value=name)  # Function
    # Column C (Status) and columns E onward (formulas) intentionally left untouched.
    all_detail.cell(row=row, column=4, value=hours)  # Development man-hours (h)

all_detail["C33"] = "Sample skill requirement description."
all_detail["C34"] = "Sample skill requirement description."

# --- Business Flow(system admin): the embedded diagram depicts the
# real project's actual admin-side navigation/feature structure (a
# genuine business-flow design, not generic sample content) -- removed
# entirely for a public sample rather than left in, since it cannot be
# meaningfully "replaced with sample data" (it's a picture, not a
# value) and keeping it would risk exposing real system design.
# Worksheet itself, its layout/formatting, and everything else about
# it is untouched.
business_flow = wb["Business Flow(system admin)"]
business_flow._images = []
business_flow._charts = []

wb.save(DEST)
print(f"Saved sanitized sample template: {DEST}")
