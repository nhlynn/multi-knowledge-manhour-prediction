"""One-off script that generates ``import/ssd/ssd_import_template.xlsx``
from ``simple_resource/ssd_import_export_format.xlsx``.

Not part of the application's runtime import/export code paths (see
``services/team_template_validator.py`` / ``services/ssd_import_parser.py``
for those) -- this is a standalone, one-time sanitization script, kept
alongside its output for reproducibility/traceability (e.g. if the real
source template's structure ever changes and this public sample needs
regenerating). Never overwrites ``SOURCE`` -- only ever reads it and
writes to ``DEST``.

Purpose: the file this produces is what ``routes/upload.py::download_template``
serves to an SSD user (via the ``sample_template_path`` in
``utils/migrations/ssd_import_export_config.py``), and what
``services/team_template_validator.py`` accepts an uploaded copy of. It
must therefore keep the workbook's STRUCTURE identical to the real
template -- worksheet names, column order, the three-row detail header,
merged cells, formulas, borders, row heights, column widths -- while
replacing every real-project-identifying VALUE with generic sample
text/round numbers so no real customer data is ever exposed publicly.

What gets sanitized:
- Both detail sheets (``詳細設計～システムテスト 本番移行`` and its ``_2``
  variant): each real task row's 機能名/機能概要/要件/見積根拠 text,
  難易度, and 調整工数 (N-Q) values are replaced with generic samples.
  Real task rows are located the same way ``services/ssd_import_parser.py``
  locates them (non-blank 機能名 that isn't a category header or a
  repeated field-header), so this stays in lockstep with the parser
  rather than hardcoding row numbers. Category header rows
  (``1.処理・画面`` etc.) and the header rows themselves are structural
  and left untouched. The standard-hours (J-M) and estimate (R-U)
  columns are VLOOKUP/=J+N formulas and are never written -- only the
  literal adjustment inputs are.
- ``見積総額``'s sample project title (B2 -- a literal string, not a
  formula on this sheet).
- ``対応方針`` and ``前提条件``: real project narrative, prerequisites,
  and any embedded code snippets are cleared, leaving each sheet's own
  header/label structure intact so the download still shows a user
  where that content goes.

What is deliberately LEFT AS-IS:
- ``難易度別標準工数`` -- the S/A/B/C/D standard-hours lookup table is
  reference data showing HOW to fill difficulty-based hours, not
  project-identifying data; a realistic example there helps whoever
  fills the template. (The parser reads it as a lookup, never as
  knowledge.)
- Every other sheet's structural labels, headers, and formulas.

Known, unavoidable characteristic (an openpyxl limitation, not
something this script introduces): any workbook re-saved via openpyxl
loses the *cached* display value of every formula cell workbook-wide.
The formulas themselves are preserved verbatim, and opening this file
in real Microsoft Excel recalculates and displays correct values
immediately. This doesn't affect the app's own import parsing --
``services/ssd_import_parser.py`` reads the lookup sheet and the
literal adjustment columns directly, resolving standard/estimate hours
itself rather than trusting any formula's cached value.
"""

import openpyxl
from openpyxl.cell.cell import MergedCell

SOURCE = "simple_resource/ssd_import_export_format.xlsx"
DEST = "import/ssd/ssd_import_template.xlsx"

# Mirror services/ssd_import_parser.py's structural constants so this
# script locates task rows exactly the way the parser does.
_DETAIL_SHEETS = [
    "詳細設計～システムテスト 本番移行",
    "詳細設計～システムテスト 本番移行_2",
]
_FIELD_HEADER_ROW = 5
_DATA_START_ROW = 8
_FUNC_NAME_COL = 3       # C 機能名
_OVERVIEW_COL = 4        # D 機能概要
_REQUIREMENT_COL = 5     # E 要件（ユースケース）
_DIFFICULTY_COL = 6      # F 難易度
_BASIS_COL = 8           # H 見積根拠
_WORK_EXPL_COL = 9       # I 工数説明（難易度説明）— free text, project-specific
_ADJUSTMENT_COLS = [14, 15, 16, 17]  # N-Q 調整工数

_HEADER_LABELS = {"機能名", "機能概要", "要件（ユースケース）", "難易度",
                  "新規/改定", "見積根拠", "No", "No."}

# Rotating generic values so the sanitized sample still shows a variety
# of difficulties and a couple of adjustment examples.
_SAMPLE_DIFFICULTIES = ["A", "B", "C"]
_SAMPLE_ADJUSTMENTS = [{}, {17: -1}, {16: 1}]  # occasional Q/P tweaks


def _is_category_header(text: str) -> bool:
    head = text.strip().split(".", 1)[0]
    return bool(head) and head.isdigit()


def _is_repeated_header(name: str) -> bool:
    return name in _HEADER_LABELS


def _set(ws, row, col, value):
    """Write only if the target isn't a merged non-anchor cell."""
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        return
    cell.value = value


def _sanitize_detail_sheet(ws) -> None:
    sample_n = 0
    for row in range(_DATA_START_ROW, ws.max_row + 1):
        name_val = ws.cell(row=row, column=_FUNC_NAME_COL).value
        name = str(name_val).strip() if name_val is not None else ""
        if not name:
            continue
        if _is_category_header(name) or _is_repeated_header(name):
            continue

        sample_n += 1
        _set(ws, row, _FUNC_NAME_COL, f"Sample Function {sample_n}")
        # Only replace descriptive cells that actually held real text,
        # so a blank stays blank (matches the real template's shape).
        if ws.cell(row=row, column=_OVERVIEW_COL).value is not None:
            _set(ws, row, _OVERVIEW_COL, f"Sample overview {sample_n}")
        if ws.cell(row=row, column=_REQUIREMENT_COL).value is not None:
            _set(ws, row, _REQUIREMENT_COL, f"Sample requirement {sample_n}")
        if ws.cell(row=row, column=_BASIS_COL).value is not None:
            _set(ws, row, _BASIS_COL, f"Sample estimate basis {sample_n}")
        if ws.cell(row=row, column=_WORK_EXPL_COL).value is not None:
            _set(ws, row, _WORK_EXPL_COL, f"Sample work explanation {sample_n}")

        # Difficulty: only overwrite a row that actually had one (some
        # 全般/移行 rows legitimately leave it blank).
        if ws.cell(row=row, column=_DIFFICULTY_COL).value is not None:
            _set(ws, row, _DIFFICULTY_COL,
                 _SAMPLE_DIFFICULTIES[(sample_n - 1) % len(_SAMPLE_DIFFICULTIES)])

        # Clear every real adjustment first, then apply a small generic
        # sample set -- never touches the J-M / R-U formula columns.
        for col in _ADJUSTMENT_COLS:
            if ws.cell(row=row, column=col).value is not None:
                _set(ws, row, col, None)
        for col, val in _SAMPLE_ADJUSTMENTS[(sample_n - 1) % len(_SAMPLE_ADJUSTMENTS)].items():
            _set(ws, row, col, val)


def _clear_narrative_sheet(ws, header_rows_to_keep: int) -> None:
    """Blank out every cell below ``header_rows_to_keep`` that isn't a
    formula, leaving the sheet's own header/label structure so the
    download still shows a user where content goes."""
    for row in range(header_rows_to_keep + 1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            if isinstance(cell, MergedCell):
                continue
            if isinstance(cell.value, str) and cell.value.startswith("="):
                continue  # preserve any formula (e.g. title header)
            cell.value = None


def main() -> None:
    wb = openpyxl.load_workbook(SOURCE)

    for sheet_name in _DETAIL_SHEETS:
        if sheet_name in wb.sheetnames:
            _sanitize_detail_sheet(wb[sheet_name])

    # Project-specific free text on the roll-up sheet: title (B2),
    # last-updated date (H3), and the two proposal-approach headings
    # (B4 / B18). All literal strings here; the C/D/F/H hour/amount
    # cells are formulas and left untouched.
    if "見積総額" in wb.sheetnames:
        summary = wb["見積総額"]
        _set(summary, 2, 2, "Sample Project")          # B2 title
        _set(summary, 3, 8, "最終更新：YYYY/MM/DD")       # H3 last-updated
        _set(summary, 4, 2, "① Sample approach 1")      # B4 proposal ① heading
        _set(summary, 18, 2, "② Sample approach 2")     # B18 proposal ② heading

    # Free-form project narrative / prerequisites / reference sheets:
    # clear real content, keep each sheet's own label structure (top
    # rows: title + header labels) so the download still shows a user
    # where content goes. These sheets are never read for knowledge
    # (only the detail + lookup sheets are — see
    # services/ssd_import_parser.py), so clearing them costs nothing and
    # guarantees no real project text, person names (体制), code
    # snippets (対応方針), or dates leak into the public sample.
    #
    # header_rows_to_keep is per-sheet: enough to preserve that sheet's
    # own column-header row (below which everything is project content).
    _NARRATIVE_SHEETS = {
        "対応方針": 4,
        "前提条件": 4,
        "体制": 4,       # keeps 会社/お名前/役割/備考 header, drops real names
        "成果物": 4,
        "スケジュール": 4,
        "システム構成図": 2,
    }
    for sheet_name, keep in _NARRATIVE_SHEETS.items():
        if sheet_name in wb.sheetnames:
            _clear_narrative_sheet(wb[sheet_name], header_rows_to_keep=keep)

    wb.save(DEST)
    print(f"Wrote sanitized SSD template: {DEST}")


if __name__ == "__main__":
    main()
