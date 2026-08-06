"""One-off script that generated ``import/kikan/kikan_import_template.xlsx``
from ``simple_resource/kikan_import_export_template.xlsx``.

Not part of the application's runtime import/export code paths (see
``services/team_template_validator.py`` / ``services/kikan_export_builder.py``
for those) -- this is a standalone, one-time sanitization script, kept
alongside its output for reproducibility/traceability (e.g. if the real
source template's structure ever changes and this sample needs
regenerating).

Replaces every real-project-identifying value in ``工数詳細`` (the
7 real function rows: category, function/screen ID, base hours) with
generic sample text or round sample numbers, while leaving the
worksheet's structure, column names/order, data types, formulas, data
validation, merged cells, formatting, row heights, and column widths
completely untouched -- only specific cells' *values* are replaced.

``工数詳細``'s ``機能名称`` (function name) column is a live
``=VLOOKUP(...)`` formula reading from ``機能一覧`` -- rather than
overwrite it with a literal (which would silently turn a formula cell
into a plain value, unlike every other cell this script touches), the
matching ``機能一覧`` cells the formula actually depends on (its
ScreenID and function-name columns, for exactly the 7 rows this join
touches) are updated to the same generic sample identifiers/names, so
the formula keeps working and correctly displays sample text. No other
``機能一覧`` cell (its own ``業務分類``/``番号``/``機能ID``/``内容``
columns) is touched here -- sanitizing the rest of ``機能一覧`` is a
separate, later step.

Known, unavoidable characteristic (an openpyxl limitation, not
something this script introduces -- see
``import/bamawl/build_sample_template.py``'s own note): any workbook
re-saved via openpyxl loses the *cached* display value of every
formula cell workbook-wide, regardless of which cells were actually
edited. The formulas themselves are preserved verbatim, and opening
this file in real Microsoft Excel recalculates and displays correct
sample values immediately (Excel's default calculation mode is
Automatic) -- this only matters for a tool that reads cached values
without recalculating (e.g. pandas) before the file has ever been
opened/saved in Excel.
"""

import openpyxl

SOURCE = "simple_resource/kikan_import_export_template.xlsx"
DEST = "import/kikan/kikan_import_template.xlsx"

wb = openpyxl.load_workbook(SOURCE)

# --- 7 generic sample functions -- same identifiers/names used in
# both 工数詳細 (the join *source*, column C) and 機能一覧 (the join
# *target*, columns D/E), so 工数詳細!機能名称's VLOOKUP formula keeps
# resolving correctly to clean sample text instead of #N/A. ---
SAMPLE_SCREEN_IDS = [f"SCR-SAMPLE-{i:03d}" for i in range(1, 8)]
SAMPLE_FUNCTION_NAMES = [f"Sample Function {i}" for i in range(1, 8)]
SAMPLE_HOURS = [40, 30, 50, 40, 30, 40, 30]

# --- 工数詳細: category (merged block's top-left cell), the join key,
# and base hours -- 機能名称 (D) and every phase/total formula are left
# completely untouched. ---
detail = wb["工数詳細"]
detail["A5"] = "Sample Category"
for i, row in enumerate(range(5, 12)):
    detail.cell(row=row, column=3, value=SAMPLE_SCREEN_IDS[i])  # 機能ID (join key)
    detail.cell(row=row, column=6, value=SAMPLE_HOURS[i])  # 実装工数(h)

# --- 機能一覧: only the two columns 工数詳細's VLOOKUP formula
# actually reads (ScreenID, 機能名称), for exactly the 7 rows the join
# touches -- everything else in this sheet (業務分類/番号/機能ID/内容)
# is untouched, sanitized separately later. ---
function_list = wb["機能一覧"]
for i, row in enumerate(range(2, 9)):
    function_list.cell(row=row, column=4, value=SAMPLE_SCREEN_IDS[i])  # ScreenID
    function_list.cell(row=row, column=5, value=SAMPLE_FUNCTION_NAMES[i])  # 機能名称

wb.save(DEST)
