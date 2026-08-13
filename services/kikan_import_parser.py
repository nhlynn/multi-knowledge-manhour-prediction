"""KiKan Team's own thin import-enrichment wrapper.

Unlike ``services/sgl_import_parser.py`` (which reads its sheet
directly because SGL's two-row header genuinely can't be expressed
through the generic single-header-row ``column_mapping`` engine),
KiKan Team's own ``工数詳細`` sheet has an ordinary single header row
and is already fully, correctly parsed by the generic
``services.excel_parser.excel_to_nested_json`` -- including its own
``Status`` and ``機能ID`` columns, captured via that engine's own
generic ``extra_columns`` mechanism (see
``utils/migrations/kikan_import_export_config.py``'s
``KIKAN_IMPORT_COLUMN_MAPPING["extra_columns"]``).

The ONE thing the generic, single-sheet engine genuinely can't reach:
KiKan's workbook has a SECOND worksheet, ``機能一覧`` (Function List),
holding each function's OWN ``機能ID`` and ``内容`` (Description) --
different columns from anything on ``工数詳細``, joined to it only via
that sheet's ``ScreenID`` column matching ``工数詳細``'s own
"機能ID"-labeled column (see
``services/kikan_export_builder.py``'s own "機能一覧 sync" module note
for the full picture of this relationship, which originally lived as a
live ``=VLOOKUP(...)`` formula in the pristine template).

This module's only job is that one cross-sheet join: call the generic
engine for everything else, then enrich each task with ``機能一覧``'s
own ``function_id``/``content`` fields by matching on the ``screen_id``
the generic engine's ``extra_columns`` mechanism already captured.
Both new fields flow through to the final JSON/search output for free,
via the same generic-field passthrough ``services/sgl_import_parser.py``'s
own ``work_detail``/``block`` fields already rely on.
"""

import logging
from typing import Any

import openpyxl

from services.excel_parser import _find_column, excel_to_nested_json

logger = logging.getLogger(__name__)

_FUNCTION_LIST_SHEET = "機能一覧"
_FUNCTION_LIST_HEADER_ROW = 1


def kikan_excel_to_nested_json(excel_path: str) -> list[dict[str, Any]]:
    """Parse KiKan Team's workbook via the generic engine, then enrich
    every task with ``機能一覧``'s own ``機能ID``/``内容`` fields (as
    ``function_id``/``content``) by matching that sheet's ``ScreenID``
    column against the task's own ``screen_id`` (captured from
    ``工数詳細``'s "機能ID"-labeled column by the generic engine's
    ``extra_columns`` mechanism -- see
    ``utils/migrations/kikan_import_export_config.py``).

    A task whose ``screen_id`` doesn't match any ``機能一覧`` row (or
    has none at all) is left exactly as the generic engine produced it
    -- ``function_id``/``content`` are simply never added for it, same
    as any other optional field this pipeline doesn't force onto every
    task.
    """
    from utils.migrations.kikan_import_export_config import KIKAN_IMPORT_COLUMN_MAPPING

    nested_json = excel_to_nested_json(excel_path, column_mapping=KIKAN_IMPORT_COLUMN_MAPPING)

    screen_id_map = _read_function_list(excel_path)
    if not screen_id_map:
        return nested_json

    matched = 0
    for cat in nested_json:
        for task in cat.get("tasks", []):
            screen_id = task.get("screen_id")
            extra = screen_id_map.get(screen_id) if screen_id else None
            if not extra:
                continue
            if extra.get("function_id"):
                task["function_id"] = extra["function_id"]
            if extra.get("content"):
                task["content"] = extra["content"]
            matched += 1

    logger.info(
        "KiKan import: matched %d task(s) to '%s' rows via ScreenID.",
        matched, _FUNCTION_LIST_SHEET,
    )
    return nested_json


def _read_function_list(excel_path: str) -> dict[str, dict[str, str]]:
    """Read ``機能一覧`` directly (via openpyxl, not the generic
    pandas-based engine -- this one small, single-purpose sheet isn't
    worth its own column_mapping/DataFrame machinery) into
    ``{screen_id: {"function_id": ..., "content": ...}}``.

    Returns an empty dict (logged, not raised) if the sheet or its
    expected columns are missing -- a workbook that's otherwise a
    valid KiKan import shouldn't fail entirely just because this one
    enrichment step couldn't run; every task simply keeps going without
    ``function_id``/``content``, same as before this module existed.
    """
    try:
        wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    except (OSError, KeyError, ValueError):
        logger.exception("KiKan import: couldn't open %s to read '%s'.", excel_path, _FUNCTION_LIST_SHEET)
        return {}

    try:
        if _FUNCTION_LIST_SHEET not in wb.sheetnames:
            logger.warning(
                "KiKan import: '%s' worksheet not found; function_id/content "
                "enrichment skipped.", _FUNCTION_LIST_SHEET,
            )
            return {}
        ws = wb[_FUNCTION_LIST_SHEET]

        header_to_col: dict[str, int] = {}
        for cell in ws[_FUNCTION_LIST_HEADER_ROW]:
            if cell.value is not None:
                header_to_col[str(cell.value).strip()] = cell.column

        screen_id_col = header_to_col.get(_find_column(list(header_to_col), "ScreenID") or "")
        function_id_col = header_to_col.get(_find_column(list(header_to_col), "機能ID") or "")
        content_col = header_to_col.get(_find_column(list(header_to_col), "内容") or "")
        if not screen_id_col:
            logger.warning(
                "KiKan import: '%s' worksheet has no ScreenID column; function_id/content "
                "enrichment skipped.", _FUNCTION_LIST_SHEET,
            )
            return {}

        result: dict[str, dict[str, str]] = {}
        for row in ws.iter_rows(min_row=_FUNCTION_LIST_HEADER_ROW + 1):
            screen_id_val = row[screen_id_col - 1].value
            screen_id = str(screen_id_val).strip() if screen_id_val is not None else ""
            if not screen_id:
                continue
            entry: dict[str, str] = {}
            if function_id_col:
                fid_val = row[function_id_col - 1].value
                if fid_val is not None and str(fid_val).strip():
                    entry["function_id"] = str(fid_val).strip()
            if content_col:
                content_val = row[content_col - 1].value
                if content_val is not None and str(content_val).strip():
                    entry["content"] = str(content_val).strip()
            if entry:
                result[screen_id] = entry

        return result
    finally:
        # CRITICAL on Windows: openpyxl's read_only mode keeps the
        # workbook's underlying zip file handle open for its lazy row
        # iteration, and never closes it on its own. Without this,
        # every KiKan import leaves excel_path locked at the OS level
        # for as long as this process keeps running -- on Windows,
        # that surfaces later as "[WinError 32] The process cannot
        # access the file because it is being used by another process"
        # the next time someone tries to delete or re-upload that exact
        # knowledge-base file. finally guarantees this runs on every
        # exit path above, not just the normal one.
        wb.close()