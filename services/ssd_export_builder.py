"""SSD Team's own Export Builder.

Independent from the Bamawl/KiKan/SGL builders — no shared code beyond
the generic helpers already in ``services/excel_parser.py`` and the
shared ``BaseExportService`` interface. Built on top of SSD Team's real
internal Excel workbook,
``simple_resource/ssd_import_export_format.xlsx`` — every export copies
that workbook, populates the copy, and saves the copy to the export
path; the real template is never modified, and the sanitized public
sample (``import/ssd/ssd_import_template.xlsx``) is never read here.

Design — why SSD needs its own builder:

- SSD's "詳細設計～システムテスト 本番移行" worksheet has a three-row header
  and, per phase, THREE hour columns (標準作業工数 J–M, 調整工数 N–Q,
  見積工数 R–U) rather than one — see ``services/ssd_import_parser.py``.
  Export writes all three groups back so the exported file shows the
  full breakdown the user saw imported.
- Task rows sit under section-number category headers (``1.処理・画面``,
  ``3.全般`` …), interleaved with 小計 subtotal rows. Writable task rows
  are discovered from the template itself (rows carrying the per-row
  見積工数 ``=標準+調整`` formula or a 標準作業工数 VLOOKUP), never
  hardcoded, so a template edit that adds/moves rows is picked up.
- Each selected task is written under a row belonging to ITS OWN
  category (the category label recorded on the task), so a re-exported
  全般 task can never land under 処理・画面 purely by list ordering —
  mirroring how ``services/sgl_export_builder.py`` keeps tasks within
  their own block.

Values, not formulas: this builder writes each phase's 標準/調整/見積 as
literal numbers (taken from the task's own per-phase breakdown captured
at import — see ``standard_hours``/``adjustment_hours``/``estimate_hours``
on each activity). It deliberately does NOT rely on re-triggering the
template's VLOOKUP/``=標準+調整`` formulas, because an unused row cleared
of its 難易度 would otherwise leave those formulas showing ``#N/A``.
Writing literals keeps every row clean and correct whether used or not,
and the 見積総額 roll-up sheet's own SUM formulas (which read the R–U
見積工数 column) recompute correctly from those literals.

Only the primary detail sheet's task rows and the 見積総額 title cell are
written. The second proposal sheet ("…_2"), the 難易度別標準工数 lookup,
and every narrative/reference sheet are left exactly as the template
has them (per the export design: import reads both proposals, export
writes the primary one).
"""

import logging
import os
import re
from datetime import date
from typing import Any

import openpyxl
from openpyxl.cell.cell import MergedCell

from services.base_export_service import BaseExportService, ExportContext

logger = logging.getLogger(__name__)

# Japanese weekday abbreviations, Monday-first (date.weekday(): Mon=0).
_JP_WEEKDAYS = ["月", "火", "水", "木", "金", "土", "日"]

# The 最終更新 (last-updated) caption on each sheet, e.g.
# "最終更新：2/27/2026（金）". On export its date is refreshed to today.
# Matched loosely on the "最終更新" prefix so the stale template date and
# weekday are fully replaced regardless of their old value/format.
_LAST_UPDATED_RE = re.compile(r"^\s*最終更新")


def _today_last_updated() -> str:
    """"最終更新：M/D/YYYY（<曜>）" for today, matching the template's
    own caption format (M/D/YYYY plus a Japanese weekday)."""
    today = date.today()
    wd = _JP_WEEKDAYS[today.weekday()]
    return f"最終更新：{today.month}/{today.day}/{today.year}（{wd}）"


def _refresh_last_updated(wb) -> int:
    """Replace every sheet's 最終更新 caption with today's date. Returns
    the number of cells updated. Skips merged non-anchor cells."""
    stamp = _today_last_updated()
    updated = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and _LAST_UPDATED_RE.match(cell.value):
                    if not isinstance(cell, MergedCell):
                        cell.value = stamp
                        updated += 1
    return updated

# Primary detail sheet written on export. The "…_2" second-proposal
# sheet is intentionally left as the template has it.
SSD_DETAIL_SHEET = "詳細設計～システムテスト 本番移行"
SSD_SUMMARY_SHEET = "見積総額"
SSD_SUMMARY_TITLE_CELL = "B2"

# Three-row header (mirrors services/ssd_import_parser.py).
_FIELD_HEADER_ROW = 5
_PHASE_LABEL_ROW = 7
_DATA_START_ROW = 8

_FUNC_NAME_HEADER = "機能名"
_FUNC_OVERVIEW_HEADER = "機能概要"
_REQUIREMENT_HEADER = "要件（ユースケース）"
_DIFFICULTY_HEADER = "難易度"
_KIND_HEADER = "新規/改定"
_BASIS_HEADER = "見積根拠"

_STANDARD_GROUP_HEADER = "標準作業工数（人日）"
_ADJUSTMENT_GROUP_HEADER = "調整工数（人日）"
_ESTIMATE_GROUP_HEADER = "見積工数（人日）"

_PHASE_LABELS = ["詳細設計", "実装", "単体テスト", "結合テスト"]

# A genuine task row carries a per-row 見積工数 formula ("=J9+N9",
# "=2/8", literal, etc.) or a 標準作業工数 VLOOKUP. Category-header rows
# and 小計 subtotal rows do not, so this distinguishes real task slots
# without hardcoding row numbers. Matched against the first 見積工数
# column (R) and first 標準作業工数 column (J).
_VLOOKUP_RE = re.compile(r"VLOOKUP", re.IGNORECASE)

# A 小計 subtotal row's 見積 cell is a SUM over a multi-row RANGE
# ("=SUM(R8:R25)") — never a per-row task formula. Excluded from
# writable task rows so its subtotal formula is never cleared or
# overwritten.
_SUM_RANGE_RE = re.compile(r"^=SUM\([A-Za-z]+\d+:[A-Za-z]+\d+\)$", re.IGNORECASE)


class SsdExportError(ValueError):
    """Raised when SSD Team's export template can't be built from, or
    the system data doesn't fit it -- see ``build_ssd_workbook``."""


def _resolve_field_columns(ws) -> dict[str, int]:
    return {
        str(cell.value).strip(): cell.column
        for cell in ws[_FIELD_HEADER_ROW]
        if cell.value is not None and str(cell.value).strip()
    }


def _resolve_work_note_column(ws, basis_col: int | None) -> int | None:
    """Return the 工数説明 column — the second column of the 見積根拠
    merged header (H5:I5). Located from that merge span, mirroring
    services/ssd_import_parser.py; None if 見積根拠 isn't merged across
    two columns."""
    if not basis_col:
        return None
    for merged in ws.merged_cells.ranges:
        if (merged.min_row == _FIELD_HEADER_ROW
                and merged.min_col == basis_col
                and merged.max_col > basis_col):
            return merged.max_col
    return None


def _resolve_group_columns(ws, field_columns: dict[str, int], group_header: str) -> list[tuple[str, int]]:
    """``[(phase_label, column_index), ...]`` for one merged phase group,
    read from that group-label cell's merge span and each column's row-7
    phase label — never hardcoded."""
    group_col = field_columns.get(group_header)
    if not group_col:
        return []
    end_col = group_col
    for merged in ws.merged_cells.ranges:
        if merged.min_row == _FIELD_HEADER_ROW and merged.min_col == group_col:
            end_col = merged.max_col
            break
    columns: list[tuple[str, int]] = []
    for col in range(group_col, end_col + 1):
        label = ws.cell(row=_PHASE_LABEL_ROW, column=col).value
        if label is not None and str(label).strip():
            columns.append((str(label).strip(), col))
    return columns


def _is_category_header(text: str) -> bool:
    head = text.strip().split(".", 1)[0]
    return bool(head) and head.isdigit()


def _discover_task_rows_by_category(
    ws, func_name_col: int, first_std_col: int | None, first_est_col: int | None,
) -> dict[str, list[int]]:
    """Discover every writable task row, grouped by the category header
    it sits under. A row is a writable task slot if it carries a per-row
    見積工数 formula/value or a 標準作業工数 VLOOKUP -- category headers and
    小計 subtotal rows have neither. Category is forward-filled from the
    most recent section-number header (``1.処理・画面`` etc.).
    """
    by_category: dict[str, list[int]] = {}
    current_category = ""

    for row in range(_DATA_START_ROW, ws.max_row + 1):
        name_val = ws.cell(row=row, column=func_name_col).value
        name = str(name_val).strip() if name_val is not None else ""

        if name and _is_category_header(name):
            current_category = name
            by_category.setdefault(current_category, [])
            continue

        if not current_category:
            continue

        # Is this a writable task slot? Check the first 標準 (VLOOKUP) or
        # first 見積 (per-row formula/value) column. A 小計 subtotal row
        # (見積 = "=SUM(range)") is explicitly excluded.
        writable = False
        if first_est_col is not None:
            rv = ws.cell(row=row, column=first_est_col).value
            if isinstance(rv, str) and _SUM_RANGE_RE.match(rv.replace(" ", "")):
                # Subtotal row — never writable.
                continue
        if first_std_col is not None:
            jv = ws.cell(row=row, column=first_std_col).value
            if isinstance(jv, str) and _VLOOKUP_RE.search(jv):
                writable = True
        if not writable and first_est_col is not None:
            rv = ws.cell(row=row, column=first_est_col).value
            if isinstance(rv, str) and rv.startswith("="):
                writable = True
            elif isinstance(rv, (int, float)):
                writable = True

        # A row that already holds a real 機能名 (the template's sample
        # task) is also a writable slot even if its formulas were lost.
        if not writable and name and not _is_category_header(name):
            writable = True

        if writable:
            by_category[current_category].append(row)

    return by_category


def _set(ws, row: int, col: int | None, value) -> None:
    """Write only if the target column is known and not a merged
    non-anchor cell."""
    if not col:
        return
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        return
    cell.value = value


def _clear_task_row(
    ws, row: int, cols: dict[str, int],
    std_cols: list[tuple[str, int]], adj_cols: list[tuple[str, int]], est_cols: list[tuple[str, int]],
    basis_col: int | None = None, work_note_col: int | None = None,
) -> None:
    """Blank every writable cell of a task row before any selected task
    is written, so an unused row can never retain the template's own
    sample data or a previous export's values."""
    for header in (_FUNC_NAME_HEADER, _FUNC_OVERVIEW_HEADER, _REQUIREMENT_HEADER,
                   _DIFFICULTY_HEADER, _KIND_HEADER, _BASIS_HEADER):
        _set(ws, row, cols.get(header), None)
    _set(ws, row, basis_col, None)
    _set(ws, row, work_note_col, None)
    for _, col in (*std_cols, *adj_cols, *est_cols):
        _set(ws, row, col, None)


def _num(value) -> float:
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def build_ssd_workbook(
    filepath: str,
    categories: list[dict[str, Any]],
    template_path: str,
    project_name: str | None = None,
) -> None:
    """Populate SSD Team's own Excel template with Preview's data and
    save the result to ``filepath``.

    Only the primary "詳細設計～システムテスト 本番移行" sheet's task rows and
    the 見積総額 title cell are written; every other sheet is left as the
    template has it.

    Each selected task is written under a row belonging to its own
    category, with 機能名/機能概要/要件/難易度/見積根拠 and all three phase
    hour groups (標準作業工数/調整工数/見積工数) as literal values from the
    task's own per-phase breakdown.

    Raises:
        SsdExportError: if the template file/worksheet/columns can't be
            found, or a category has more selected functions than its
            template rows can hold.
    """
    if not os.path.isfile(template_path):
        raise SsdExportError(f"SSD Team's export template file is missing: {template_path}")

    wb = openpyxl.load_workbook(template_path)
    if SSD_DETAIL_SHEET not in wb.sheetnames:
        raise SsdExportError(
            f"SSD Team's export template is missing the required '{SSD_DETAIL_SHEET}' worksheet."
        )
    ws = wb[SSD_DETAIL_SHEET]

    cols = _resolve_field_columns(ws)
    func_name_col = cols.get(_FUNC_NAME_HEADER)
    difficulty_col = cols.get(_DIFFICULTY_HEADER)
    if not (func_name_col and difficulty_col):
        raise SsdExportError(
            f"SSD Team's export template's required columns (機能名/難易度) could not be located "
            f"in '{SSD_DETAIL_SHEET}'."
        )

    std_cols = _resolve_group_columns(ws, cols, _STANDARD_GROUP_HEADER)
    adj_cols = _resolve_group_columns(ws, cols, _ADJUSTMENT_GROUP_HEADER)
    est_cols = _resolve_group_columns(ws, cols, _ESTIMATE_GROUP_HEADER)
    if not est_cols:
        raise SsdExportError(
            f"SSD Team's export template's 見積工数 phase columns could not be located "
            f"in '{SSD_DETAIL_SHEET}'."
        )

    first_std_col = std_cols[0][1] if std_cols else None
    first_est_col = est_cols[0][1] if est_cols else None

    # 見積根拠 (H) and its second merged column 工数説明 (I). Written back
    # so a re-exported row carries the full column set the original had.
    basis_col = cols.get(_BASIS_HEADER)
    work_note_col = _resolve_work_note_column(ws, basis_col)

    rows_by_category = _discover_task_rows_by_category(
        ws, func_name_col, first_std_col, first_est_col,
    )
    if not rows_by_category:
        raise SsdExportError(
            f"SSD Team's export template has no recognizable task rows in "
            f"'{SSD_DETAIL_SHEET}'; cannot populate it."
        )

    # Clear every discovered writable row up front, so unused rows never
    # keep sample/previous data.
    for rows in rows_by_category.values():
        for row in rows:
            _clear_task_row(ws, row, cols, std_cols, adj_cols, est_cols, basis_col, work_note_col)

    std_by_label = {label: col for label, col in std_cols}
    adj_by_label = {label: col for label, col in adj_cols}
    est_by_label = {label: col for label, col in est_cols}

    # Assign each selected task to a row in ITS OWN category's pool.
    pools = {cat: list(rows) for cat, rows in rows_by_category.items()}

    for cat in categories:
        cat_name = cat.get("category", "")
        pool = pools.get(cat_name)
        for task in cat.get("tasks", []):
            if pool is None or not pool:
                raise SsdExportError(
                    f"Category '{cat_name or '(unknown)'}' has more selected functions than "
                    f"SSD Team's export template has rows for it — reduce the selection, or "
                    f"update the template."
                )
            row = pool.pop(0)
            _write_task(
                ws, row, task, cols,
                std_by_label, adj_by_label, est_by_label,
                basis_col, work_note_col,
            )

    # 見積総額 project title (a literal string cell; its hour/amount cells
    # are formulas that recompute from the detail sheet's 見積工数 column).
    if project_name and SSD_SUMMARY_SHEET in wb.sheetnames:
        summary = wb[SSD_SUMMARY_SHEET]
        title_cell = summary[SSD_SUMMARY_TITLE_CELL]
        if not isinstance(title_cell, MergedCell):
            title_cell.value = project_name

    # Stamp every sheet's 最終更新 caption with today's date, so the
    # exported file reflects when it was actually produced rather than
    # the template's stale hardcoded date.
    _refresh_last_updated(wb)

    wb.save(filepath)
    logger.info(
        "SSD export written to %s (%d categories) from template %s",
        filepath, len(categories), template_path,
    )


def _write_task(
    ws, row: int, task: dict[str, Any], cols: dict[str, int],
    std_by_label: dict[str, int], adj_by_label: dict[str, int], est_by_label: dict[str, int],
    basis_col: int | None = None, work_note_col: int | None = None,
) -> None:
    """Write one task into ``row``: its descriptive fields plus all
    three phase hour groups as literal values from the task's own
    per-phase activity breakdown."""
    _set(ws, row, cols.get(_FUNC_NAME_HEADER), task.get("task", ""))
    _set(ws, row, cols.get(_FUNC_OVERVIEW_HEADER), task.get("overview", "") or None)
    _set(ws, row, cols.get(_REQUIREMENT_HEADER), task.get("requirement", "") or None)
    _set(ws, row, cols.get(_DIFFICULTY_HEADER), task.get("difficulty", "") or None)
    _set(ws, row, cols.get(_KIND_HEADER), task.get("kind", "") or None)
    # 見積根拠 and 工数説明 (I column) — the estimate-basis and hours-note
    # free text, so a re-exported row carries the full column set the
    # original had, not just name/hours.
    _set(ws, row, basis_col, task.get("basis", "") or None)
    _set(ws, row, work_note_col, task.get("work_note", "") or None)

    # Per-phase hours from the task's activities. The Preview payload
    # calls this list "activities" (built by search_service); the raw
    # parser mapping calls it "task_details" — accept either so the
    # builder works whether fed a Preview selection or mapping data.
    activity_list = task.get("activities") or task.get("task_details") or []
    by_phase = {a.get("task_detail", ""): a for a in activity_list}
    for label in _PHASE_LABELS:
        act = by_phase.get(label)
        if act is None:
            continue
        _set(ws, row, std_by_label.get(label), _num(act.get("standard_hours")))
        _set(ws, row, adj_by_label.get(label), _num(act.get("adjustment_hours")))
        _set(ws, row, est_by_label.get(label), _num(act.get("estimate_hours")))


class SsdExportBuilder(BaseExportService):
    """SSD Team's export strategy — copies SSD's real internal template
    and writes the Preview selection into the primary detail sheet with
    the full 標準/調整/見積 per-phase breakdown. See ``build_ssd_workbook``.
    """

    team_name = "SSD Team"

    @staticmethod
    def template_path(app_root_path: str) -> str:
        """Path to SSD Team's real internal Excel template — the export
        base. Import validation and Template Download instead use the
        sanitized public copy (import/ssd/ssd_import_template.xlsx);
        export always copies the real workbook, never the sanitized one.
        """
        return os.path.join(app_root_path, "simple_resource", "ssd_import_export_format.xlsx")

    @staticmethod
    def fixed_phase_labels(app_root_path: str) -> list[str]:
        """The template's four fixed phase labels (詳細設計/実装/単体テスト/
        結合テスト), read from the 見積工数 group's sub-columns. The Preview
        page fixes each SSD task's activity rows to exactly this set, so
        a non-matching activity (which would have no column to export
        into) can never be added. Returns [] if the template can't be
        read."""
        path = SsdExportBuilder.template_path(app_root_path)
        if not os.path.isfile(path):
            return []
        try:
            wb = openpyxl.load_workbook(path)
            if SSD_DETAIL_SHEET not in wb.sheetnames:
                return []
            ws = wb[SSD_DETAIL_SHEET]
            cols = _resolve_field_columns(ws)
            return [label for label, _ in _resolve_group_columns(ws, cols, _ESTIMATE_GROUP_HEADER)]
        except Exception:
            return []

    def build(self, context: ExportContext) -> None:
        build_ssd_workbook(
            context.filepath,
            context.categories,
            template_path=context.template_path,
            project_name=context.project_name,
        )