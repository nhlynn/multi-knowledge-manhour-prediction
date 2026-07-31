"""Reads a previously exported workbook back into a display-friendly structure.

Moved out of ``routes/export.py`` so the route stays thin. This is the
read-only, display-side mirror of ``services/export_workbook_service.py``
(which writes the same row layout this module reads back) and of
``services/remark_html.py`` (whose HTML-to-rich-text conversion this
module's rich-text-to-HTML conversion undoes for viewing). No behavior
changed from the original ``routes/export.py::_read_export_detail`` /
``_fill_to_hex`` / ``_rich_text_to_html``.
"""

import re
from html import escape

from openpyxl import load_workbook
from openpyxl.cell.rich_text import CellRichText


def read_export_detail(filepath) -> dict:
    """Read an exported workbook back into a display-friendly structure.

    ``filepath`` may be a path string (local disk) or a file-like object
    such as ``io.BytesIO`` (downloaded from GCS) — ``openpyxl.load_workbook``
    accepts either.

    Mirrors the exact row layout ``export_workbook_service.build_workbook``
    writes (title, Created By/Date, headers, category/task rows, Total,
    Remark), so this only works for files this app generated itself —
    which is the only kind that ever lands in the exports folder.
    """
    wb = load_workbook(filepath, rich_text=True)
    try:
        ws = wb.active

        title = (ws["A1"].value or "").strip()
        project_name = re.sub(r"\s*Manhour\s*$", "", title).strip() or title
        created_by = ws["E3"].value or ""
        date_value = ws["E4"].value
        date_str = date_value.strftime("%Y-%m-%d") if hasattr(date_value, "strftime") else (date_value or "")

        categories = []
        total_row = None
        r = 6
        while True:
            cat_value = ws.cell(row=r, column=1).value
            task_value = ws.cell(row=r, column=2).value

            if isinstance(cat_value, str) and cat_value.strip() == "Total":
                total_row = r
                break
            if cat_value is None and task_value is None:
                # No more data rows (shouldn't normally happen before a
                # Total row, but avoids an infinite loop on a malformed file).
                break

            if cat_value:
                categories.append({"name": cat_value, "rows": []})
            if not categories:
                categories.append({"name": "", "rows": []})

            estimate = ws.cell(row=r, column=3).value
            working_day = round(estimate / 8, 2) if isinstance(estimate, (int, float)) else ""
            categories[-1]["rows"].append({
                "task": task_value or "",
                "estimate": estimate,
                "working_day": working_day,
                "remarks": ws.cell(row=r, column=5).value or "",
            })
            r += 1

        grand_total = ws.cell(row=total_row, column=3).value if total_row else 0
        grand_working_day = round(grand_total / 8, 2) if isinstance(grand_total, (int, float)) else ""

        remark_row = (total_row + 3) if total_row else None
        remark_cell = ws.cell(row=remark_row, column=1) if remark_row else None
        remark_html = rich_text_to_html(remark_cell.value if remark_cell else None)
        remark_bg = fill_to_hex(remark_cell.fill) if remark_cell else None
        remark_hyperlink = remark_cell.hyperlink.target if (remark_cell and remark_cell.hyperlink) else None

        return {
            "project_name": project_name,
            "created_by": created_by,
            "date_str": date_str,
            "categories": categories,
            "grand_total": grand_total,
            "grand_working_day": grand_working_day,
            "remark_html": remark_html,
            "remark_bg": remark_bg,
            "remark_hyperlink": remark_hyperlink,
        }
    finally:
        wb.close()


def fill_to_hex(fill) -> str | None:
    """Convert an openpyxl cell fill to a CSS hex color, or None if unfilled."""
    if not fill or fill.fill_type != "solid" or not fill.fgColor:
        return None
    rgb = fill.fgColor.rgb
    if isinstance(rgb, str) and len(rgb) == 8:
        return "#" + rgb[2:]
    return None


def rich_text_to_html(value) -> str:
    """Convert a cell's rich-text (or plain string) value into safe HTML.

    Only for read-only display of our own generated files — reconstructs
    bold/italic/underline/font color and line breaks from the openpyxl
    rich-text runs. This is the display-side mirror of
    ``services/remark_html.py``'s HTML-to-rich-text conversion.
    """
    if value in (None, "") or (isinstance(value, str) and value.strip() == "No remark added."):
        return '<span class="text-muted fst-italic">No remark added.</span>'

    if isinstance(value, str):
        return escape(value).replace("\n", "<br>")

    runs = value if isinstance(value, CellRichText) else [value]
    html_parts = []
    for part in runs:
        text = part.text if hasattr(part, "text") else str(part)
        escaped = escape(text).replace("\n", "<br>")
        font = getattr(part, "font", None)
        if font is None:
            html_parts.append(escaped)
            continue

        color_hex = None
        if font.color and isinstance(font.color.rgb, str) and len(font.color.rgb) == 8:
            color_hex = "#" + font.color.rgb[2:]

        open_tags, close_tags = "", ""
        if font.b:
            open_tags += "<strong>"
            close_tags = "</strong>" + close_tags
        if font.i:
            open_tags += "<em>"
            close_tags = "</em>" + close_tags
        if font.u:
            open_tags += "<u>"
            close_tags = "</u>" + close_tags

        style_attr = f' style="color: {color_hex};"' if color_hex else ""
        html_parts.append(f"<span{style_attr}>{open_tags}{escaped}{close_tags}</span>")

    return "".join(html_parts)
