"""Bamawl Team's own Export Builder.

Unlike the generic export path (``services/export_workbook_service.py``,
which builds a fresh workbook from scratch via a column-layout config),
Bamawl Team's export is built directly on top of Bamawl Team's single
official Excel workbook (``simple_resource/bamawl_import_export_format_filled.xlsx``)
-- the same file (identical structure -- same 7 worksheets, same
``ALL_Detail`` layout) used on the import side by
``services/team_template_validator.py``. There is deliberately no separate
import-only or export-only template file:

- The template workbook is loaded as-is and saved back out — every
  worksheet (``ReqDefinition``, ``FunctionList``, ``TotalManhour``,
  ``ALL_Detail``, ``Infra Manhour``, ``Business Flow(system admin)``,
  ``Milestone``) is preserved, under its original name, with its
  original formatting (fonts, borders, merges, column widths) untouched
  except for the specific cells this module writes into.
- Only the ``ALL_Detail`` worksheet's data rows are populated, from
  MHES's internal Category → Task → Activity data (the same
  ``categories`` structure the Preview page sends to
  ``routes/export.py::export_excel``) — reusing Bamawl Team's
  configured ``column_mapping`` (phases mode; see
  ``utils/migrations/bamawl_import_export_config.py``) to know which
  column each phase (Development, Code Review, ...) belongs in, the
  export-side mirror of how ``services/team_template_validator.py`` reads
  that same sheet on the way in.
- The ``FunctionList`` worksheet's "Function Name" column is
  regenerated from that exact same task set (one row per ``ALL_Detail``
  task, same "No." numbering) instead of being left as the template's
  original sample function list — before this, every Bamawl export
  shipped that fixed sample list regardless of which functions the user
  had actually selected during the estimation workflow (chatbot
  selection -> Preview), unrelated to the real, correctly-scoped
  ``ALL_Detail`` data next to it. No other worksheet reads or depends
  on ``FunctionList``'s contents (checked directly -- no cross-sheet
  formula references it), so regenerating it has no effect elsewhere.
- ``ReqDefinition``'s title cell (``B1``, merged ``B1:C1`` -- the
  template's original sample text, e.g. "Bamawl HR & Attendance
  System") is replaced with the Preview page's Project Name field,
  exactly as the user typed it (no trimming/casing/reformatting), or
  left blank if that field was empty. Only this one cell's *value* is
  touched -- the merge and formatting are untouched.
- ``ReqDefinition``'s Purpose / In Scope / Document Specifications /
  Functional Specifications / Server Environment / Out of Scope value
  cells are always cleared, deliberately -- never filled from AI
  generation, the database, or the Preview payload (there is no
  MHES data model field these could even come from) -- so a user fills
  them in manually after export. Only the value cell next to each
  label is cleared; the label text, its formatting, and its merged
  range are untouched.
- ``ALL_Detail``'s ``Status`` column is written from a task's own
  ``status`` field when present -- captured at import time from
  ``ALL_Detail``'s own ``Status`` column (see
  ``utils/migrations/bamawl_import_export_config.py``'s
  ``extra_columns``) and carried through Preview/search generically,
  the same mechanism KiKan's own ``Status`` column and SGL's own
  ``work_detail`` field use. Left blank only for a task with no such
  value (e.g. a brand-new function added directly in Preview, never
  imported from a workbook).
- ``TotalManhour``'s data rows are entirely rewritten from
  ``TOTAL_MANHOUR_GROUPS`` -- each row's value computed as the sum,
  across this export's tasks, of a set of ``ALL_Detail`` phase labels
  (Development, Code Review, Requirement definition, Basic overall
  design, Basic design, Unit Test, Combined Test, Comprehensive Test,
  Test Data Creation, User Manual, Accidental Work, Risk, Management
  Manhours) -- see ``_populate_total_manhour``. This replaces both the
  old template-baked placeholder numbers (Infra setup, Deployment,
  Training, Operations Support, Maintenance, Device Test -- none of
  which had any corresponding MHES field, and never changed between
  exports) and ``Development``'s old ``=ALL_Detail!AD15`` fixed-row
  formula (which could silently mismatch if a project's task count
  differed from the template's own built-in sample range -- see the
  "Known limitation" note above). The トータル(hr/1person)/(days/
  1person)/(month/1person) block's own formulas are untouched.
- ``Business Flow(system admin)``'s content -- confirmed to be exactly
  one embedded picture, no shapes/SmartArt/text boxes/connectors --
  is always stripped from the export. The worksheet itself is kept
  (an empty sheet, not removed from the workbook); only its picture
  and drawing part are dropped on save.

**Known limitation** (a direct consequence of reusing this specific
template file rather than building a fresh one): ``ALL_Detail`` has its
own internal subtotal formulas (e.g. ``=SUM(D5:D14)``) immediately below
the task rows, and ``TotalManhour`` references ``ALL_Detail`` by a fixed
row number (e.g. ``=ALL_Detail!AD15``) -- both calibrated to the
template's own built-in row range. This module writes into the existing
task-row block only (never touching or shifting the subtotal
rows/formulas below it) and raises ``BamawlExportError`` rather than
overflow into them if a project has more tasks than that block holds.
Within capacity, per-task numbers are always correct; the *subtotal*
rows/‐other sheets' formulas were not re-derived for an arbitrary task
count and may not exactly match if the count differs from the
template's own built-in sample.
"""

import logging
import os
from copy import copy
from typing import Any

import openpyxl
from openpyxl.formula.translate import Translator
from openpyxl.utils import get_column_letter

from services.base_export_service import BaseExportService, ExportContext
from services.excel_parser import _find_column, _normalize_header, _safe_float

logger = logging.getLogger(__name__)

# FunctionList isn't part of Bamawl Team's import column_mapping (it's
# an export-only, informational sheet in the template -- see module
# docstring), so its sheet/column layout is fixed here instead.
# ALL_Detail's row-2 ratio/coefficient cells the derived-phase formulas
# multiply by (e.g. =E5*F$2). Edited percentages are written here on
# export so the formulas recompute with them.
COEFFICIENT_ROW = 2

FUNCTION_LIST_SHEET = "FunctionList"
FUNCTION_LIST_HEADER_ROW = 1
FUNCTION_LIST_NO_COLUMN = 2  # "No."
FUNCTION_LIST_NAME_COLUMN = 3  # "Function Name"

# Same reasoning as FunctionList above: ReqDefinition isn't part of the
# import column_mapping either, so its title cell's location is fixed
# here.
REQ_DEFINITION_SHEET = "ReqDefinition"
REQ_DEFINITION_TITLE_CELL = "B1"

# These sections' VALUE cells (column C, next to their label in column
# B) are intentionally always left blank on export -- see
# _blank_req_definition_sections. Matched by label text (not hardcoded
# row numbers), since the label itself is what identifies each section.
REQ_DEFINITION_LABEL_COLUMN = 2  # B
REQ_DEFINITION_VALUE_COLUMN = 3  # C
REQ_DEFINITION_BLANK_SECTION_LABELS = [
    "Purpose",
    "In Scope",
    "Document Specifications",
    "Functional Specifications",
    "Server Environment",
    "Out of Scope",
]

# The template's "Business Flow(system admin)" sheet carries no cell
# content at all (dims A1:A1) -- its only content is one embedded
# picture (a diagram/screenshot) covering the sheet, confirmed directly
# against the drawing XML: exactly one <xdr:pic> anchor, no shapes,
# connectors, or SmartArt. Stripped entirely on export -- see
# _strip_business_flow_content.
BUSINESS_FLOW_SHEET = "Business Flow(system admin)"

# TotalManhour's own row/column layout -- see _populate_total_manhour.
TOTAL_MANHOUR_SHEET = "TotalManhour"
TOTAL_MANHOUR_LABEL_COLUMN = 2  # B
TOTAL_MANHOUR_VALUE_COLUMN = 3  # C
TOTAL_MANHOUR_TOTAL_LABEL_PREFIX = "トータル"

# Each row of TotalManhour is now one of these groups -- its value is
# the sum, across every task in the export, of the listed
# ALL_Detail/phase_columns labels (the same "label" strings configured
# in BAMAWL_IMPORT_COLUMN_MAPPING["phase_columns"], and the same ones
# _phase_value already matches an activity's task_detail against when
# writing ALL_Detail's own per-task phase columns). Order here is the
# row order written into the sheet.
#
# Every one of BAMAWL_IMPORT_COLUMN_MAPPING's 26 phase labels is
# accounted for in exactly one group below -- deliberately, so this
# breakdown's grand total (the トータル block, which just sums whatever
# is in these rows) never silently drops part of ALL_Detail's data.
# "Test Data Creation" has its own row for this reason even though it
# wasn't named as one of the original requested categories -- flag to
# a human if it should instead be folded into an existing group (e.g.
# Unit Test or Combined Test) or removed.
TOTAL_MANHOUR_GROUPS: list[tuple[str, list[str]]] = [
    ("Development", ["Development"]),
    ("Code Review", ["Code Review"]),
    ("Requirement definition", ["Prototype", "Prototype Review"]),
    ("Basic overall design", [
        "Business Flow", "Business Flow Review",
        "ERD", "ERD Review", "DFD", "DFD Review",
        "DB Design", "DB Design Review",
    ]),
    ("Basic design", ["Screen/Form/Function", "Screen/Form/Function Review"]),
    ("Unit Test", [
        "Unit Test Specification", "Unit Test Review", "Unit Test Implementation",
    ]),
    ("Combined Test", [
        "Combined Test Specification", "Combined Test Review", "Combined Test Implementation",
    ]),
    ("Comprehensive Test", ["Comprehensive Test Implementation"]),
    ("Test Data Creation", ["Test Data Creation"]),
    ("User Manual", ["User Manual"]),
    ("Accidental Work", ["Accidental Work"]),
    ("Risk", ["Risk"]),
    ("Management Manhours", ["Management Manhours"]),
]

# ALL_Detail's "Status" column (see BAMAWL_ALL_DETAIL_HEADERS in
# utils/migrations/bamawl_import_export_config.py) isn't part of
# column_mapping's phase_columns -- it's captured on import via that
# same config's "extra_columns" (a task's Status cell -> task["status"],
# same mechanism KiKan's own "Status" column uses) and written back out
# here by column NAME (resolved through _resolve_template_columns, not
# hardcoded), so a task with a status value round-trips through
# export exactly like KiKan's does. A task with no status (e.g. one
# added directly in Preview rather than imported from a workbook) is
# simply left blank -- same reasoning Bamawl's export already blanks
# ReqDefinition's free-text sections.
STATUS_COLUMN_NAME = "Status"


class BamawlExportError(ValueError):
    """Raised when Bamawl Team's export template can't be built from,
    or the system data doesn't fit it -- see ``build_bamawl_workbook``."""


def _resolve_template_columns(ws, header_row: int) -> dict[str, int]:
    """Map each header cell's column name -- reproducing the same
    duplicate-suffixing (``"Review(h)"``/``"Review(h).1"``) and
    whitespace-stripping ``services/excel_parser.py``'s import path
    applies when it reads this same sheet via pandas -- to its real
    1-indexed column position in the template, so ``column_mapping``'s
    configured names (already written in that pandas-equivalent form)
    resolve back to the actual cells to write into.
    """
    seen: dict[str, int] = {}
    name_to_col: dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        raw = ws.cell(row=header_row, column=c).value
        raw_text = "" if raw is None else str(raw)
        n = seen.get(raw_text, 0)
        seen[raw_text] = n + 1
        pandas_name = raw_text if n == 0 else f"{raw_text}.{n}"
        name_to_col[pandas_name.strip()] = c
    return name_to_col


def _column_index(name_to_col: dict[str, int], target: str | None) -> int | None:
    """Resolve a configured column name to its column index, tolerant
    of whitespace/case the same way the import side is."""
    if not target:
        return None
    matched = _find_column(list(name_to_col.keys()), target)
    return name_to_col.get(matched) if matched else None


def _template_capacity(
    ws, data_start_row: int, task_col: int, phase_cols: list[tuple[str, int]],
) -> int:
    """Return how many task rows are available below the header before
    the template's own subtotal/summary block starts.

    Scans down from ``data_start_row``: a row with a real task name is
    counted; a row with a blank task name is still counted (a blank
    separator row within the sample block) *unless* it already has a
    value in one of the phase columns -- that's the signature of a
    subtotal row (e.g. ``=SUM(D5:D14)``), which marks the boundary.
    """
    row = data_start_row
    while row <= ws.max_row:
        task_val = ws.cell(row=row, column=task_col).value
        if not task_val:
            phase_has_value = any(
                ws.cell(row=row, column=col_idx).value not in (None, "")
                for _, col_idx in phase_cols
            )
            if phase_has_value:
                break
        row += 1
    return row - data_start_row


def _phase_value(activities: list[dict[str, Any]], label: str) -> float:
    """Return the estimate_hours of the activity matching ``label``
    (whitespace/case-insensitive), or 0.0 if this task has none."""
    target = _normalize_header(label)
    for act in activities:
        if _normalize_header(act.get("task_detail") or "") == target:
            return _safe_float(act.get("estimate_hours"))
    return 0.0


def _populate_function_list(wb, task_names: list[str]) -> None:
    """Regenerate ``FunctionList``'s "Function Name" column from
    exactly the given task names (one row per name, numbered 1..N) —
    the user's selected functions, nothing more.

    Missing or extra sample rows are handled without disturbing
    formatting: rows within the template's original sample range keep
    their existing cell style as-is (only the value changes); any row
    needed beyond that range copies the style of the last original
    sample row so a project with more functions than the template's
    original sample count still looks consistent. Unused rows within
    the original range are left blank (value cleared, formatting kept).

    A missing ``FunctionList`` worksheet is not fatal — logged and
    skipped, since it's an informational sheet rather than the
    required ``ALL_Detail`` data (see ``build_bamawl_workbook``).
    """
    if FUNCTION_LIST_SHEET not in wb.sheetnames:
        logger.warning(
            "Bamawl Team's export template has no '%s' worksheet; skipping "
            "Function Name list population.", FUNCTION_LIST_SHEET,
        )
        return

    ws = wb[FUNCTION_LIST_SHEET]
    data_start_row = FUNCTION_LIST_HEADER_ROW + 1
    original_last_row = ws.max_row

    for r in range(data_start_row, original_last_row + 1):
        ws.cell(row=r, column=FUNCTION_LIST_NO_COLUMN).value = None
        ws.cell(row=r, column=FUNCTION_LIST_NAME_COLUMN).value = None

    style_row = max(original_last_row, data_start_row)
    style_no = ws.cell(row=style_row, column=FUNCTION_LIST_NO_COLUMN)
    style_name = ws.cell(row=style_row, column=FUNCTION_LIST_NAME_COLUMN)

    for i, name in enumerate(task_names, start=1):
        row = data_start_row + i - 1
        no_cell = ws.cell(row=row, column=FUNCTION_LIST_NO_COLUMN, value=i)
        name_cell = ws.cell(row=row, column=FUNCTION_LIST_NAME_COLUMN, value=name)
        if row > original_last_row:
            no_cell.font, no_cell.alignment, no_cell.border = (
                copy(style_no.font), copy(style_no.alignment), copy(style_no.border),
            )
            name_cell.font, name_cell.alignment, name_cell.border = (
                copy(style_name.font), copy(style_name.alignment), copy(style_name.border),
            )

    logger.info(
        "Populated '%s' with %d selected function(s) (template originally had %d sample rows).",
        FUNCTION_LIST_SHEET, len(task_names), max(original_last_row - data_start_row + 1, 0),
    )


def _populate_req_definition_title(wb, project_name: str | None) -> None:
    """Replace ``ReqDefinition``'s title cell with the Preview page's
    Project Name, used exactly as entered (no trimming/casing changes),
    or blank if it was empty/None.

    Only this one cell's value is set — its existing formatting (font,
    fill, borders) and its merge (``B1:C1``) are untouched, since
    writing into the top-left cell of an existing merged range is all
    that's needed; no other cell in this worksheet is touched.

    A missing ``ReqDefinition`` worksheet is not fatal — logged and
    skipped, same reasoning as a missing ``FunctionList``.
    """
    if REQ_DEFINITION_SHEET not in wb.sheetnames:
        logger.warning(
            "Bamawl Team's export template has no '%s' worksheet; skipping "
            "project title replacement.", REQ_DEFINITION_SHEET,
        )
        return

    ws = wb[REQ_DEFINITION_SHEET]
    ws[REQ_DEFINITION_TITLE_CELL].value = project_name or None
    logger.info(
        "Set '%s'!%s to Project Name %r.",
        REQ_DEFINITION_SHEET, REQ_DEFINITION_TITLE_CELL, project_name or "",
    )


def _blank_req_definition_sections(wb) -> None:
    """Clear the value cell next to each of
    ``REQ_DEFINITION_BLANK_SECTION_LABELS`` (Purpose, In Scope, Document
    Specifications, Functional Specifications, Server Environment, Out
    of Scope) — deliberately, on every export, regardless of what the
    template's own sample content or any other data source might say.

    Nothing populates these from AI generation, the database, or the
    Preview payload — they're simply set to blank so a user fills them
    in manually after export. Only the value cell (column C) is
    touched; the label itself (column B) and everything else in the
    worksheet (including the just-written title in ``B1``) are
    untouched.

    Matched by the label text actually present in column B for each
    row (case/whitespace-tolerant via ``_normalize_header``), not a
    hardcoded row number — a label merged across several rows (e.g.
    "Server Environment" spanning B7:B11, "Out of Scope" spanning
    B12:B19) only has its label text in the merge's first row, which is
    exactly the row whose value cell (the corresponding merged range's
    first cell, e.g. C7 or C12) needs clearing.
    """
    if REQ_DEFINITION_SHEET not in wb.sheetnames:
        logger.warning(
            "Bamawl Team's export template has no '%s' worksheet; skipping "
            "section blanking.", REQ_DEFINITION_SHEET,
        )
        return

    ws = wb[REQ_DEFINITION_SHEET]
    targets = {_normalize_header(label) for label in REQ_DEFINITION_BLANK_SECTION_LABELS}
    blanked: list[str] = []

    for r in range(1, ws.max_row + 1):
        label = ws.cell(row=r, column=REQ_DEFINITION_LABEL_COLUMN).value
        if label and _normalize_header(str(label)) in targets:
            ws.cell(row=r, column=REQ_DEFINITION_VALUE_COLUMN).value = None
            blanked.append(str(label))

    missing = targets - {_normalize_header(b) for b in blanked}
    if missing:
        logger.warning(
            "Bamawl export: could not find these '%s' section label(s) to blank: %s",
            REQ_DEFINITION_SHEET, sorted(missing),
        )
    logger.info("Blanked %d '%s' section value(s): %s", len(blanked), REQ_DEFINITION_SHEET, blanked)


def _populate_total_manhour(wb, tasks: list[dict[str, Any]]) -> None:
    """Rewrite every ``TotalManhour`` data row from ``TOTAL_MANHOUR_GROUPS``,
    each row's value computed as the sum, across every task in this
    export, of its listed phase labels' ``estimate_hours`` (via the
    same ``_phase_value`` helper used to write ``ALL_Detail``'s own
    per-task phase columns) -- so every row is now genuinely derived
    from this export's own data, replacing the old template-baked
    placeholder numbers (Infra setup, Deployment, Training, Operations
    Support, Maintenance, Device Test) that had no corresponding MHES
    field and never changed between exports.

    This also replaces ``Development``'s old ``=ALL_Detail!AD15``
    formula with a plain computed value -- summed directly here rather
    than referencing a fixed row in ``ALL_Detail``, so it no longer
    depends on the current project's task count matching the
    template's own built-in sample-row range (see this module's
    "Known limitation" note above ``build_bamawl_workbook``, which no
    longer applies to this sheet).

    Every label (column B) and value (column C) in the sheet's old
    data-row block is cleared first, then rewritten one row per
    ``TOTAL_MANHOUR_GROUPS`` entry starting at row 2 -- so no leftover
    old label (e.g. a since-removed placeholder row) can survive past
    however many groups are configured. The trailing トータル
    (hr/1person)/(days/1person)/(month/1person) block's own
    ``SUM``/division formulas are left completely untouched; they
    continue to total whatever is now in the rewritten rows.

    Raises:
        BamawlExportError: if ``TOTAL_MANHOUR_GROUPS`` has more rows
            than the sheet has room for before its トータル block.
    """
    if TOTAL_MANHOUR_SHEET not in wb.sheetnames:
        logger.warning(
            "Bamawl Team's export template has no '%s' worksheet; skipping "
            "TotalManhour breakdown.", TOTAL_MANHOUR_SHEET,
        )
        return

    ws = wb[TOTAL_MANHOUR_SHEET]

    data_start_row = 2
    total_row = None
    for r in range(data_start_row, ws.max_row + 1):
        label = ws.cell(row=r, column=TOTAL_MANHOUR_LABEL_COLUMN).value
        if label and str(label).strip().startswith(TOTAL_MANHOUR_TOTAL_LABEL_PREFIX):
            total_row = r
            break

    if total_row is None:
        logger.warning(
            "Bamawl export: could not locate the '%s' total block in '%s'; "
            "skipping TotalManhour breakdown to avoid guessing at the wrong rows.",
            TOTAL_MANHOUR_TOTAL_LABEL_PREFIX, TOTAL_MANHOUR_SHEET,
        )
        return

    capacity = total_row - data_start_row
    if len(TOTAL_MANHOUR_GROUPS) > capacity:
        raise BamawlExportError(
            f"TOTAL_MANHOUR_GROUPS has {len(TOTAL_MANHOUR_GROUPS)} row(s), but "
            f"'{TOTAL_MANHOUR_SHEET}' only has room for {capacity} before its "
            f"'{TOTAL_MANHOUR_TOTAL_LABEL_PREFIX}' total block."
        )

    # Clear the sheet's entire old data-row block (label + value) first,
    # so no leftover label/number from a previous layout survives past
    # however many groups are written back below.
    for r in range(data_start_row, total_row):
        ws.cell(row=r, column=TOTAL_MANHOUR_LABEL_COLUMN).value = None
        ws.cell(row=r, column=TOTAL_MANHOUR_VALUE_COLUMN).value = None

    all_activities = [task.get("activities", []) or [] for task in tasks]
    for i, (group_label, phase_labels) in enumerate(TOTAL_MANHOUR_GROUPS):
        row = data_start_row + i
        total = sum(
            _phase_value(activities, phase_label)
            for activities in all_activities
            for phase_label in phase_labels
        )
        ws.cell(row=row, column=TOTAL_MANHOUR_LABEL_COLUMN, value=group_label)
        ws.cell(row=row, column=TOTAL_MANHOUR_VALUE_COLUMN, value=total)

    logger.info(
        "Rewrote '%s' with %d group row(s), computed from %d task(s)' ALL_Detail "
        "phase data.", TOTAL_MANHOUR_SHEET, len(TOTAL_MANHOUR_GROUPS), len(tasks),
    )


def _strip_business_flow_content(wb) -> None:
    """Remove every editable business-flow element from the
    ``Business Flow(system admin)`` worksheet -- diagrams, flowcharts,
    SmartArt, shapes, text boxes, connector lines, and images copied
    from the Bamawl Team template -- while keeping the worksheet
    itself (an empty sheet, not removed from the workbook).

    ``ws._images`` covers the one embedded picture the template
    actually has here (confirmed directly against the sheet's drawing
    XML: a single ``<xdr:pic>`` anchor, nothing else) -- clearing it
    drops that image, and with it the sheet's entire drawing part, when
    the workbook is saved. ``openpyxl`` does not model
    shapes/SmartArt/text boxes/connectors as separate objects it could
    round-trip in the first place (it only understands cells, images,
    and charts), so loading this workbook through it already discards
    any such elements before this function ever runs; clearing
    ``_images`` (and, defensively, ``_charts``, in case a future
    template revision adds one) is what actually removes the one kind
    of visual content this library *does* otherwise preserve.

    A missing worksheet is not fatal — logged and skipped, same
    reasoning as a missing ``FunctionList``/``ReqDefinition``.
    """
    if BUSINESS_FLOW_SHEET not in wb.sheetnames:
        logger.warning(
            "Bamawl Team's export template has no '%s' worksheet; nothing to strip.",
            BUSINESS_FLOW_SHEET,
        )
        return

    ws = wb[BUSINESS_FLOW_SHEET]
    image_count = len(ws._images)
    chart_count = len(ws._charts)
    ws._images = []
    ws._charts = []
    logger.info(
        "Stripped '%s': removed %d image(s) and %d chart(s); worksheet itself kept.",
        BUSINESS_FLOW_SHEET, image_count, chart_count,
    )


def build_bamawl_workbook(
    filepath: str,
    categories: list[dict[str, Any]],
    column_mapping: dict[str, Any],
    template_path: str,
    project_name: str | None = None,
    phase_coefficients: list[dict[str, Any]] | None = None,
) -> None:
    """Populate Bamawl Team's own Excel template with system data and
    save it to ``filepath``.

    Args:
        filepath: Where to save the populated workbook.
        categories: The exported Category → Task → Activity data (same
            shape ``services/export_workbook_service.py`` receives).
            Every task across every category is written, in order, as
            one ``ALL_Detail`` row -- Bamawl Team's sheet has no
            category column of its own (see
            ``BAMAWL_IMPORT_COLUMN_MAPPING``'s fixed literal
            ``category``), so category boundaries aren't represented.
        column_mapping: Bamawl Team's configured phases-mode column
            mapping (``sheet``, ``header_row``, ``task_column``,
            ``id_column``, ``phase_columns``, ``total_column``).
        template_path: Path to Bamawl Team's single official template
            workbook (``simple_resource/bamawl_import_export_format_filled.xlsx``
            -- see ``routes/export.py::_bamawl_template_path``), the
            same file used on the import side.
        project_name: The Preview page's Project Name field, used
            verbatim to replace ``ReqDefinition``'s title cell. None or
            empty leaves that cell blank (see
            ``_populate_req_definition_title``).

    Raises:
        BamawlExportError: if the template file/worksheet/columns
            can't be found, or the project has more tasks than the
            template's task-row block can hold without touching its
            subtotal rows (see module docstring).
    """
    if not os.path.isfile(template_path):
        raise BamawlExportError(f"Bamawl Team's export template file is missing: {template_path}")

    sheet_name = column_mapping.get("sheet")
    header_row = column_mapping.get("header_row") or 1

    wb = openpyxl.load_workbook(template_path)

    if sheet_name not in wb.sheetnames:
        raise BamawlExportError(
            f"Bamawl Team's export template is missing the required '{sheet_name}' worksheet."
        )
    ws = wb[sheet_name]

    name_to_col = _resolve_template_columns(ws, header_row)
    task_col = _column_index(name_to_col, column_mapping.get("task_column"))
    if task_col is None:
        raise BamawlExportError(
            "Bamawl Team's export template's task column could not be located; "
            "cannot populate ALL_Detail."
        )
    id_col = _column_index(name_to_col, column_mapping.get("id_column"))
    total_col = _column_index(name_to_col, column_mapping.get("total_column"))
    status_col = _column_index(name_to_col, STATUS_COLUMN_NAME)
    # Requirements column: on import it's read as each task's Category
    # (category_column), so on export write the task's category back into
    # it. None when the template has no such column (older layout).
    category_col = _column_index(name_to_col, column_mapping.get("category_column"))
    phase_cols = [
        (phase["label"], _column_index(name_to_col, phase["column"]))
        for phase in column_mapping.get("phase_columns", [])
    ]
    phase_cols = [(label, idx) for label, idx in phase_cols if idx is not None]

    # Flatten tasks in category order, carrying each task's category so
    # it can be written back into the Requirements column.
    tasks = []
    task_categories = []
    for cat in categories:
        cat_name = cat.get("category", "") or ""
        for task in cat.get("tasks", []):
            tasks.append(task)
            task_categories.append(cat_name)

    data_start_row = header_row + 1
    capacity = _template_capacity(ws, data_start_row, task_col, phase_cols)
    if len(tasks) > capacity:
        raise BamawlExportError(
            f"This project has {len(tasks)} task(s), but Bamawl Team's export template's "
            f"'{sheet_name}' worksheet only has room for {capacity} before its built-in "
            f"subtotal rows -- reduce the number of tasks, or update the template."
        )

    # Bamawl's phases are all DERIVED from Development man-hours by the
    # template's own ratio formulas (e.g. E=D*E$2, chained G=F*G$2 /
    # O=N*O$2, Risk=(D+H+J+...)*AB$2, Management=SUM(D:Z)*AC$2), and so
    # is Total(h) (=SUM(D:AC)). Preview now makes Development the only
    # editable phase and computes the rest with those same ratios, so a
    # template formula and Preview's number always agree. To keep the
    # exported workbook's LIVE auto-calculation (change Development in
    # Excel -> everything recomputes), we write only Development as a
    # literal and re-inject each derived phase's + Total's original
    # template formula, row-shifted, on every populated row. Capture
    # those formulas from the template's first data row BEFORE clearing.
    base_col = next(
        (idx for label, idx in phase_cols
         if _normalize_header(label) == _normalize_header("Development")),
        None,
    )
    formula_cols = [idx for _label, idx in phase_cols if idx != base_col]
    if total_col:
        formula_cols.append(total_col)
    # Capture each column's template formula and the row it came from.
    # Prefer the SECOND data row (data_start_row + 1): this template's
    # very first data row uses relative coefficient refs (e.g. "=D5*E2")
    # while every row below uses the correct absolute form ("=D6*E$2"),
    # and only an absolute coefficient ref survives being row-shifted to
    # other rows. Fall back to the first data row per-column if the
    # second row has no formula there.
    ref_row = data_start_row + 1 if capacity > 1 else data_start_row
    template_formulas = {}  # col -> (formula, origin_row)
    for c in formula_cols:
        v = ws.cell(row=ref_row, column=c).value
        origin = ref_row
        if not (isinstance(v, str) and v.startswith("=")):
            v = ws.cell(row=data_start_row, column=c).value
            origin = data_start_row
        if isinstance(v, str) and v.startswith("="):
            template_formulas[c] = (v, origin)

    # Apply the user's edited percentages (from Preview) to the
    # template's coefficient row (row 2), matched by phase label. The
    # derived-phase formulas reference these cells (e.g. =E5*F$2), so the
    # exported workbook recomputes with the adjusted percentages instead
    # of the template's originals — and stays live (change Development in
    # Excel → everything follows the edited %). Written before the task
    # loop and the 0%-column hiding below, so both see the edited values.
    coef_by_label = {
        _normalize_header(c.get("label")): c.get("coef")
        for c in (phase_coefficients or [])
        if c.get("label") is not None and c.get("coef") is not None
    }
    if coef_by_label:
        for label, col_idx in phase_cols:
            if col_idx == base_col:
                continue
            coef = coef_by_label.get(_normalize_header(label))
            if coef is not None:
                ws.cell(row=COEFFICIENT_ROW, column=col_idx, value=coef)

    # Clear the template's own sample rows across the whole task-row
    # block first, so no leftover sample values (or their formulas)
    # linger past however many real rows are written below.
    for r in range(data_start_row, data_start_row + capacity):
        for c in range(1, ws.max_column + 1):
            ws.cell(row=r, column=c).value = None

    unmatched_labels: set[str] = set()
    row = data_start_row
    for i, task in enumerate(tasks, start=1):
        activities = task.get("activities", []) or []
        configured_labels = {_normalize_header(label) for label, _ in phase_cols}
        for act in activities:
            norm = _normalize_header(act.get("task_detail") or "")
            if norm and norm not in configured_labels:
                unmatched_labels.add(act.get("task_detail"))

        if id_col:
            ws.cell(row=row, column=id_col, value=i)
        ws.cell(row=row, column=task_col, value=task.get("task", ""))

        # Write the task's category back into the Requirements column
        # (the same column it was read from as category_column on import).
        if category_col:
            ws.cell(row=row, column=category_col, value=task_categories[i - 1])

        if status_col:
            status = task.get("status")
            if status:
                ws.cell(row=row, column=status_col, value=status)

        # Development (base) is written as a literal; every derived
        # phase column gets its template ratio formula, translated to
        # this row (relative refs like D5 shift to D{row}; absolute
        # coefficient refs like E$2 stay put).
        if base_col is not None:
            dev_value = _phase_value(activities, "Development")
            ws.cell(row=row, column=base_col).value = dev_value or None
        for c, (formula, origin_row) in template_formulas.items():
            origin = f"{get_column_letter(c)}{origin_row}"
            dest = f"{get_column_letter(c)}{row}"
            ws.cell(row=row, column=c).value = (
                Translator(formula, origin=origin).translate_formula(dest)
            )

        row += 1

    if unmatched_labels:
        logger.warning(
            "Bamawl export: %d activity label(s) didn't match any configured phase "
            "column and were left out of '%s': %s",
            len(unmatched_labels), sheet_name, sorted(unmatched_labels),
        )

    _populate_function_list(wb, [task.get("task", "") for task in tasks])
    _populate_req_definition_title(wb, project_name)
    _blank_req_definition_sections(wb)
    _populate_total_manhour(wb, tasks)
    _strip_business_flow_content(wb)

    # Hide only the columns whose EFFECTIVE percentage is exactly 0 --
    # the edited value from Preview when the user changed it (matched by
    # phase label), otherwise the template's own coefficient. Using the
    # per-label edited value as the source of truth (not just re-reading
    # the cell) guarantees a phase the user set to a non-zero % is never
    # hidden, and one set to 0% always is -- exactly matching Preview.
    # Columns are hidden, not deleted, so every Total(h)=SUM(...) formula
    # and the TotalManhour rollup keep referencing an intact range.
    for label, col_idx in phase_cols:
        if col_idx is None or col_idx == base_col:
            continue
        eff = coef_by_label.get(_normalize_header(label))
        if eff is None:
            eff = ws.cell(row=COEFFICIENT_ROW, column=col_idx).value
        if eff == 0:
            ws.column_dimensions[get_column_letter(col_idx)].hidden = True

    wb.save(filepath)
    logger.info(
        "Built Bamawl Team export workbook: %s (%d task row(s) written into '%s')",
        filepath, len(tasks), sheet_name,
    )


class BamawlExportBuilder(BaseExportService):
    """Bamawl Team's export builder (Strategy Pattern) -- the single
    home for everything Bamawl-specific about exporting: the Strategy
    Pattern wiring (``build``), and how to resolve Bamawl Team's own
    ``column_mapping``/template path (``resolve_column_mapping``,
    ``template_path``), which used to live in ``routes/export.py`` as
    Bamawl-only helper functions. That route is now completely
    generic -- it only ever calls this class's methods, never
    Bamawl-specific logic of its own (see
    ``routes/export.py::_select_export_strategy``).

    ``build`` itself still simply delegates to ``build_bamawl_workbook``
    above (unchanged) -- this class is a thin, dedicated container
    around Bamawl's own already-existing, already-tested logic, not a
    reimplementation of it.
    """

    team_name = "Bamawl Team"

    @staticmethod
    def resolve_column_mapping(mhes_db_path: str, team_id: int) -> dict[str, Any] | None:
        """Return Bamawl Team's configured import column mapping for
        ``team_id``, or None if it hasn't been seeded yet.

        The export builder reuses this (rather than a separate
        config) — it already describes exactly which worksheet/columns
        ``ALL_Detail``'s data lives in, the same mapping
        ``services/team_template_validator.py`` reads it with.
        """
        from repositories.team_import_config_repository import TeamImportConfigRepository

        repo = TeamImportConfigRepository(mhes_db_path)
        config = repo.get_by_team_id(team_id)
        return config["column_mapping"] if config else None

    @staticmethod
    def template_path(app_root_path: str) -> str:
        """Path to Bamawl Team's Excel template (import + export share
        one format).

        Uses the git-tracked ``import/bamawl/bamawl_import_template.xlsx``
        -- the same structure Bamawl imports from, now including the
        ``Requirements`` column (read as each task's Category on import
        and written back into it on export). This replaces the old
        ``simple_resource/`` customer workbook, which (a) isn't present
        on a clean deploy and (b) predates the Requirements column, so
        an export built on it had no column to write the category into.
        """
        return os.path.join(app_root_path, "import", "bamawl", "bamawl_import_template.xlsx")

    def build(self, context: ExportContext) -> None:
        build_bamawl_workbook(
            context.filepath, context.categories, context.column_mapping,
            context.template_path, project_name=context.project_name,
            phase_coefficients=context.phase_coefficients,
        )