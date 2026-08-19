"""Builds the downloadable Excel workbook for an exported estimate.

Moved out of ``routes/export.py`` so the route stays thin — this is the
export-side counterpart to ``services/excel_parser.py`` (which turns an
uploaded Knowledge Base workbook *into* structured data; this module
turns structured Preview data *into* a workbook). No behavior changed
from the original ``routes/export.py::_build_workbook``.
"""

import logging
import re
from datetime import datetime
from html import unescape
from html.parser import HTMLParser

from openpyxl import Workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


def _css_color_to_hex(value: str) -> str | None:
    """Convert a CSS color from a style attribute ("#e60000" or
    "rgb(230, 0, 0)") to an openpyxl "RRGGBB" hex string, or None."""
    if not value:
        return None
    value = value.strip()
    m = re.search(r"#([0-9a-fA-F]{6})", value)
    if m:
        return m.group(1).upper()
    m = re.search(r"rgb\(\s*(\d+)\s*,\s*(\d+)\s*,\s*(\d+)", value)
    if m:
        r, g, b = (int(m.group(i)) for i in (1, 2, 3))
        return f"{r:02X}{g:02X}{b:02X}"
    return None


class _RemarkHtmlParser(HTMLParser):
    """Turn Preview's sanitized remark HTML into a list of formatted
    runs: (text, bold, italic, underline, color-hex-or-None). Block tags
    become newlines; list items get a bullet or running number prefix so
    the structure survives in a single Excel cell."""

    def __init__(self) -> None:
        super().__init__(convert_charrefs=True)
        self.runs: list[tuple[str, bool, bool, bool, str | None]] = []
        self._bold = 0
        self._italic = 0
        self._underline = 0
        self._colors: list[str] = []
        self._ordered_counter = 0

    def _emit(self, text: str) -> None:
        if text == "":
            return
        self.runs.append((
            text,
            self._bold > 0,
            self._italic > 0,
            self._underline > 0,
            self._colors[-1] if self._colors else None,
        ))

    def handle_starttag(self, tag, attrs):
        attrs = dict(attrs)
        if tag in ("b", "strong"):
            self._bold += 1
        elif tag in ("i", "em"):
            self._italic += 1
        elif tag == "u":
            self._underline += 1
        elif tag == "span":
            color = _css_color_to_hex(attrs.get("style", ""))
            self._colors.append(color if color else (self._colors[-1] if self._colors else None))
        elif tag in ("ol", "ul"):
            self._ordered_counter = 0
        elif tag == "br":
            self._emit("\n")
        elif tag == "li":
            if self.runs and not self.runs[-1][0].endswith("\n"):
                self._emit("\n")
            if (attrs.get("data-list") or "").lower() == "ordered":
                self._ordered_counter += 1
                self._emit("%d. " % self._ordered_counter)
            else:
                self._emit("\u2022 ")

    def handle_endtag(self, tag):
        if tag in ("b", "strong"):
            self._bold = max(0, self._bold - 1)
        elif tag in ("i", "em"):
            self._italic = max(0, self._italic - 1)
        elif tag == "u":
            self._underline = max(0, self._underline - 1)
        elif tag == "span":
            if self._colors:
                self._colors.pop()
        elif tag in ("p", "div", "li", "blockquote"):
            self._emit("\n")
        elif tag in ("ol", "ul"):
            self._ordered_counter = 0

    def handle_data(self, data):
        if data:
            self._emit(data)


def _normalize_runs(runs):
    """Collapse repeated blank lines and trim leading/trailing newlines
    across the whole run list, preserving each run's formatting."""
    # Flatten to (text, fmt) then rebuild, collapsing 3+ newlines to 2.
    text_all = "".join(r[0] for r in runs)
    text_all = re.sub(r"\n{3,}", "\n\n", text_all).strip("\n")
    if not text_all.strip():
        return []
    # Re-walk runs, dropping characters trimmed above from the head/tail.
    # Simple approach: rebuild by trimming leading/trailing newline-only
    # runs and collapsing interior blank runs.
    cleaned = []
    for text, b, i, u, color in runs:
        cleaned.append([text, b, i, u, color])
    # strip leading newline-only runs
    while cleaned and cleaned[0][0].strip("\n") == "" and "\n" in cleaned[0][0] and cleaned[0][0].strip() == "":
        cleaned.pop(0)
    while cleaned and cleaned[-1][0].strip() == "" and "\n" in cleaned[-1][0]:
        cleaned.pop()
    return [tuple(r) for r in cleaned if r[0] != ""]


def _html_to_rich_text(html: str):
    """Convert sanitized remark HTML to an openpyxl CellRichText that
    keeps bold/italic/underline/color and list structure. Returns a
    plain str when there's no formatting, or "" when empty."""
    if not html:
        return ""
    parser = _RemarkHtmlParser()
    parser.feed(html)
    runs = _normalize_runs(parser.runs)
    if not runs:
        return ""
    # If nothing is formatted, a plain string is cleaner than a
    # single-block CellRichText.
    if all(not b and not i and not u and not color for _t, b, i, u, color in runs):
        return "".join(t for t, *_ in runs).strip()
    blocks = []
    for text, b, i, u, color in runs:
        font = InlineFont(
            b=b or None,
            i=i or None,
            u="single" if u else None,
            color=color if color else None,
        )
        blocks.append(TextBlock(font, text))
    return CellRichText(blocks)


def _html_to_plain_text(html: str) -> str:
    """Flatten Preview's sanitized remark HTML to plain text (used only
    to measure how many lines the remark needs for row height)."""
    if not html:
        return ""
    parser = _RemarkHtmlParser()
    parser.feed(html)
    runs = _normalize_runs(parser.runs)
    return "".join(t for t, *_ in runs).strip()


DEFAULT_EXPORT_TEMPLATE = {
    "sheet_title": "Manhour",
    "columns": [
        {"key": "category", "label": "Category", "width": 25},
        {"key": "task", "label": "Task List", "width": 45},
        {"key": "estimate_hours", "label": "Estimate (Hours)", "width": 22},
        {"key": "working_day", "label": "Working Day", "width": 15},
        {"key": "remarks", "label": "Remarks", "width": 35},
    ],
}
"""The pre-Phase-8 column layout, reproduced exactly as data instead of
hardcoded Excel column letters — this is what every team without a
configured ``team_export_templates`` row gets, so existing exports are
byte-for-byte unaffected by Phase 8 (see docs/ARCHITECTURE.md §5h).

Recognized ``columns[].key`` values (each renders one data-table column;
unknown keys render as an empty column and are logged):
    category        -- category name, merged across that category's task rows
    task            -- numbered task name ("1. <task>")
    estimate_hours  -- the task's total hours
    working_day     -- formula: whichever "estimate_hours" column's value / 8
                       (blank if the template has no "estimate_hours" column)
    remarks         -- free-text remarks (task.get("remarks", ""))
"""


def _style_row(ws, row: int, num_cols: int, *, border=None, fill=None, font=None) -> None:
    """Apply the given border/fill/font (whichever aren't None) to every
    cell in ``row`` across ``num_cols`` columns.

    Replaces what would otherwise be several separate, near-identical
    ``for col_idx in range(1, num_cols + 1): ws.cell(...)...`` loops
    (e.g. the total row) with one helper — same cells styled with the
    same objects in the same order, just not duplicated per call site.
    """
    for col_idx in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col_idx)
        if border is not None:
            cell.border = border
        if fill is not None:
            cell.fill = fill
        if font is not None:
            cell.font = font


def build_workbook(
    filepath: str,
    project_name: str,
    created_by: str,
    categories: list,
    template_config: dict | None = None,
    project_remark: str = "",
) -> None:
    """Build an Excel workbook using a team's configured column template.

    Args:
        template_config: Optional per-team template (Phase 8 — see
            ``DEFAULT_EXPORT_TEMPLATE`` above and
            ``repositories/team_export_template_repository.py``). None
            uses ``DEFAULT_EXPORT_TEMPLATE``, reproducing the exact
            pre-Phase-8 layout.

    The title/Created-By/Date metadata rows, per-category row merging,
    and totals row are shared structure — identical for every team
    regardless of template; only the data table's columns (which ones
    appear, their order, label, and width) are configurable.
    """
    template_config = template_config or DEFAULT_EXPORT_TEMPLATE
    columns = template_config.get("columns") or DEFAULT_EXPORT_TEMPLATE["columns"]
    sheet_title = template_config.get("sheet_title") or DEFAULT_EXPORT_TEMPLATE["sheet_title"]
    num_cols = len(columns)
    col_index_by_key = {col["key"]: i for i, col in enumerate(columns, 1)}
    category_col = col_index_by_key.get("category")
    estimate_col = col_index_by_key.get("estimate_hours")
    working_day_col = col_index_by_key.get("working_day")

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title

    # --- Styles ---
    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    cat_font = Font(bold=True)
    total_font = Font(bold=True)
    total_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    center_align = Alignment(horizontal="center", vertical="center")
    wrap_align = Alignment(vertical="center", wrap_text=True)

    # Column widths
    for i, col in enumerate(columns, 1):
        if col.get("width"):
            ws.column_dimensions[get_column_letter(i)].width = col["width"]

    # --- Row 1-2: Title (merged across every configured column) ---
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=num_cols)
    title_cell = ws.cell(row=1, column=1, value=f"{project_name} {sheet_title}")
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # --- Rows 3-4: Created By / Date (label in the second-to-last
    # column, value in the last — generalizes the original D3/E3, D4/E4
    # for whatever column count this template has) ---
    label_col = max(num_cols - 1, 1)
    value_col = num_cols
    ws.cell(row=3, column=label_col, value="Created By")
    created_by_cell = ws.cell(row=3, column=value_col, value=created_by)
    created_by_cell.alignment = Alignment(horizontal="left", vertical="center")

    ws.cell(row=4, column=label_col, value="Date")
    date_cell = ws.cell(row=4, column=value_col, value=datetime.now())
    date_cell.number_format = r"yyyy\-mm\-dd"
    date_cell.alignment = Alignment(horizontal="left", vertical="center")

    # --- Row 5: Headers ---
    for i, col in enumerate(columns, 1):
        cell = ws.cell(row=5, column=i, value=col.get("label") or col["key"])
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # --- Data rows ---
    row = 6
    grand_total = 0

    for cat in categories:
        cat_start_row = row
        cat_name = cat.get("category", "")

        # Each task as a numbered row (no activity detail flattening)
        task_num = 1
        cat_total_hours = 0

        for task in cat.get("tasks", []):
            task_name = task.get("task", "")
            total_hours = task.get("total_hours", 0)

            for i, col in enumerate(columns, 1):
                key = col["key"]
                if key == "category":
                    continue  # written once per category block, after this loop
                elif key == "task":
                    cell = ws.cell(row=row, column=i, value=f"{task_num}. {task_name}")
                    cell.alignment = wrap_align
                elif key == "estimate_hours":
                    cell = ws.cell(row=row, column=i, value=total_hours)
                    cell.alignment = center_align
                elif key == "working_day":
                    value = f"={get_column_letter(estimate_col)}{row}/8" if estimate_col else None
                    cell = ws.cell(row=row, column=i, value=value)
                    cell.alignment = center_align
                elif key == "remarks":
                    cell = ws.cell(row=row, column=i, value=task.get("remarks", ""))
                    cell.alignment = wrap_align
                else:
                    logger.warning("Unknown export template column key %r; left blank.", key)
                    cell = ws.cell(row=row, column=i)
                cell.border = thin_border

            cat_total_hours += total_hours
            task_num += 1
            row += 1

        cat_end_row = row - 1
        if cat_end_row < cat_start_row:
            continue

        # Category column (merged), if this template has one
        if category_col:
            cat_row_count = cat_end_row - cat_start_row + 1
            if cat_row_count > 1:
                ws.merge_cells(
                    start_row=cat_start_row, start_column=category_col,
                    end_row=cat_end_row, end_column=category_col,
                )
            cat_cell = ws.cell(row=cat_start_row, column=category_col, value=cat_name)
            cat_cell.font = cat_font
            cat_cell.alignment = Alignment(vertical="center")
            cat_cell.border = thin_border
            for r in range(cat_start_row, cat_end_row + 1):
                ws.cell(row=r, column=category_col).border = thin_border

        grand_total += cat_total_hours

    # --- Total row ---
    total_row = row
    ws.cell(row=total_row, column=1, value="Total").font = total_font
    _style_row(ws, total_row, num_cols, border=thin_border, fill=total_fill, font=total_font)

    if estimate_col:
        ws.cell(row=total_row, column=estimate_col, value=grand_total).alignment = center_align
    if working_day_col and estimate_col:
        ws.cell(
            row=total_row, column=working_day_col,
            value=f"={get_column_letter(estimate_col)}{total_row}/8",
        ).alignment = center_align

    # --- Remark section (project-level rich-text remark, if any) ---
    # The remark arrives as sanitized HTML from Preview's rich-text
    # editor. Convert it to an openpyxl CellRichText so bold/italic/
    # underline/color and list structure survive in the cell; falls back
    # to a plain string when there's no formatting. Rendered as a
    # "Remark" label row followed by one merged, wrapped, top-aligned
    # cell spanning all columns.
    remark_rich = _html_to_rich_text(project_remark)
    if remark_rich:
        label_row = total_row + 2
        ws.cell(row=label_row, column=1, value="Remark").font = total_font
        body_row = label_row + 1
        ws.merge_cells(
            start_row=body_row, start_column=1,
            end_row=body_row, end_column=num_cols,
        )
        remark_cell = ws.cell(row=body_row, column=1, value=remark_rich)
        remark_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        # Give the merged remark row room for a few lines of text.
        line_count = _html_to_plain_text(project_remark).count("\n") + 1
        ws.row_dimensions[body_row].height = max(60, min(line_count * 15 + 10, 400))

    wb.save(filepath)