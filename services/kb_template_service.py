"""Builds the downloadable Knowledge Base import template workbook.

Moved out of ``routes/upload.py`` so the route stays thin. No behavior
changed from the original ``routes/upload.py::_build_template_workbook``.
"""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


def build_template_workbook(filepath: str) -> None:
    """Write a sample knowledge-file workbook with headers and example rows."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Man Hour"

    headers = ["Category", "Task List", "Activity Details", "Estimate (Hours)", "Buffer (Hours)"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Example rows: one task with three activities, and a second task,
    # showing that Category/Task List only need to be filled on the
    # first row of their group (merged cells or blank repeats both work,
    # since the parser forward-fills them).
    example_rows = [
        ["Sample Category 1", "Sample Task 1", "Sample Activity 1", 0.5, 2],
        ["", "", "Sample Activity 2", 0.5, ""],
        ["", "", "Sample Activity 3", 1, ""],
        ["Sample Category 2", "Sample Task 2", "Sample Activity 1", 2, 1],
        ["", "", "Sample Activity 2", 1, ""],
    ]
    start_row = 2
    last_row = start_row + len(example_rows) - 1
    for row_idx, row_data in enumerate(example_rows, start=start_row):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Thin border around every cell in the table, header included.
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows(min_row=1, max_row=last_row, min_col=1, max_col=5):
        for cell in row:
            cell.border = border

    # Merge Category / Task List / Buffer (Hours) across each task group
    # (a group starts wherever Task List is filled in), same convention
    # as the reference template.
    middle_align = Alignment(vertical="center")
    group_starts = [
        r for r in range(start_row, last_row + 1)
        if ws.cell(row=r, column=2).value not in (None, "")
    ]
    group_starts.append(last_row + 1)
    for i in range(len(group_starts) - 1):
        g_start, g_end = group_starts[i], group_starts[i + 1] - 1
        for col_idx in (1, 2, 5):
            ws.cell(row=g_start, column=col_idx).alignment = middle_align
            if g_end > g_start:
                ws.merge_cells(start_row=g_start, start_column=col_idx, end_row=g_end, end_column=col_idx)

    widths = {"A": 22, "B": 20, "C": 40, "D": 18, "E": 16}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    wb.save(filepath)
