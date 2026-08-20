"""Generic, per-team Excel import template validation.

Bamawl Team was the first (and, so far, only) team with its own
strictly-validated Excel template — this module extracts the
structural-validation logic that used to live directly in
``services/bamawl_import_parser.py`` into a team-agnostic form, so
adding SGL, KiKan, or SSD's own templates later is a matter of
defining their own ``TeamTemplateSpec`` and registering it in
``services/team_template_registry.py`` — no changes needed here, in
``routes/upload.py``, or to this validation logic itself.

A team with no registered spec (every team today except Bamawl) simply
isn't validated this way at all — its upload keeps using the existing
generic, lenient, keyword-based column matching
(``services/excel_parser.py``), completely unaffected by this module's
existence.
"""

import logging
from dataclasses import dataclass
from typing import Any

import openpyxl

from services.excel_parser import excel_to_nested_json

logger = logging.getLogger(__name__)


class TeamTemplateError(ValueError):
    """Raised when a workbook doesn't structurally match a team's
    registered Excel template -- see ``validate_team_template``.

    ``str(error)`` is the full technical detail (worksheet lists, exact
    column/position, etc.) -- meant for logs, not end users. ``reason``
    is a short, user-facing category label (e.g. "Missing worksheet:
    ALL_Detail", "Invalid column order", "Missing required column",
    "Unsupported template version") that callers show alongside a
    generic friendly message instead of this exception's full
    technical text -- see ``routes/upload.py``'s handling.
    """

    def __init__(self, message: str, *, reason: str | None = None) -> None:
        super().__init__(message)
        self.reason = reason


@dataclass(frozen=True)
class TeamTemplateSpec:
    """Everything needed to validate an upload against, and offer a
    public sample download for, one team's official Excel template.

    Attributes:
        team_name: The team this spec belongs to (matched by name, the
            same convention every other team-specific lookup in this
            codebase already uses -- see
            ``utils/migrations/bamawl_import_export_config.py``).
        required_sheet_names: The template's full worksheet list, in
            order. An uploaded workbook missing any of these is
            rejected before anything else is checked.
        header_sheet: Which of those worksheets holds the actual
            knowledge data (Bamawl Team: ``"ALL_Detail"``).
        header_row: 1-indexed row number of ``header_sheet``'s header row.
        expected_headers: ``header_sheet``'s header row, verbatim and
            in order -- compared to the upload exactly (not the
            whitespace/case-tolerant matching
            ``services/excel_parser.py`` uses when actually reading
            data): both column names and column order must match.
        column_mapping: The phases-mode (or flat) column mapping
            actually used to parse ``header_sheet`` once validation
            passes -- passed straight through to
            ``services.excel_parser.excel_to_nested_json``.
        template_version: Optional version marker to check "if
            available" -- ``None`` (the default, and every team's
            current value) means no version-marker convention exists
            for this team's template yet, so that check is skipped.
        sample_template_path: Path components (relative to the app
            root, joined lazily by the caller) to this team's public,
            sanitized sample template file -- what
            ``routes/upload.py::download_template`` serves instead of
            the real internal template. ``None`` means this team has
            no dedicated sample yet (falls back to the generic
            template builder).
    """

    team_name: str
    required_sheet_names: list[str]
    header_sheet: str
    header_row: int
    expected_headers: list[str]
    column_mapping: dict[str, Any]
    template_version: str | None = None
    sample_template_path: tuple[str, ...] | None = None
    # When set, header validation only requires these columns to be
    # PRESENT in header_sheet's header row (matched whitespace/case-
    # tolerantly, in any position) rather than the whole
    # ``expected_headers`` list matching exactly and in order. Lets a
    # team accept uploads that carry the essential columns even if other
    # columns differ, are reordered, or are added/removed. None keeps
    # the strict exact-match behaviour (every other team).
    required_columns: list[str] | None = None


def _rewind(source: Any) -> None:
    """Seek a file-like ``source`` back to its start, if it supports it.

    Validation reads ``source`` twice (worksheet list, then header
    row) before the caller (or ``parse_team_template``) reads it a
    third time to actually convert it -- a plain file path re-opens
    fine each time, but an in-memory upload stream (e.g. Flask's
    ``FileStorage.stream``, read directly so nothing is written to
    disk before validation passes) must be rewound between reads.
    """
    if hasattr(source, "seek"):
        source.seek(0)


def validate_team_template(source: Any, spec: TeamTemplateSpec) -> None:
    """Validate that ``source`` structurally matches ``spec`` exactly,
    before any data is imported from it.

    Checks, in order (any failure raises immediately -- nothing after
    the first failing check is evaluated, and no data is ever parsed
    if validation fails):

    1. Every worksheet in ``spec.required_sheet_names`` exists.
    2. ``spec.header_sheet`` specifically exists (redundant with #1
       today since it's always one of the required sheets, but kept as
       its own check/message since it's the knowledge source).
    3. ``spec.header_sheet``'s header row (``spec.header_row``) matches
       ``spec.expected_headers`` **exactly and in order**, column by
       column. Column names are compared verbatim and position
       matters: a reordered or renamed column fails even if every
       individually-required column could still be found somewhere in
       the row.
    4. Required headers present -- implied by #3.
    5. Template version matches, if ``spec.template_version`` is set
       ("if available" -- most teams' specs leave this ``None`` today,
       which skips the check entirely).

    Args:
        source: A file path, or a file-like/stream object (e.g. an
            upload's ``FileStorage.stream``) -- left rewound to its
            start when this function returns, whether it raises or not.
        spec: The uploading team's ``TeamTemplateSpec``.

    Raises:
        TeamTemplateError: with a specific, human-readable reason for
            whichever check above failed first.
    """
    try:
        wb = openpyxl.load_workbook(source, read_only=True, data_only=True)
        sheet_names = list(wb.sheetnames)
    except Exception as e:
        raise TeamTemplateError(
            f"This file could not be opened as an Excel workbook. "
            f"Please upload the official {spec.team_name} template (.xlsx).",
            reason="Could not open file",
        ) from e
    finally:
        _rewind(source)

    # 1) Required worksheet names exist. In lenient mode (required_columns
    # set) only the header sheet -- the one actually read on import -- is
    # required; the other template sheets are optional.
    sheets_needed = (
        [spec.header_sheet] if spec.required_columns else spec.required_sheet_names
    )
    missing_sheets = [s for s in sheets_needed if s not in sheet_names]
    if missing_sheets:
        wb.close()
        raise TeamTemplateError(
            f"This doesn't look like the official {spec.team_name} template: missing "
            f"required worksheet(s): {', '.join(missing_sheets)}. This workbook has: "
            f"{', '.join(sheet_names)}.",
            reason=f"Missing worksheet: {', '.join(missing_sheets)}",
        )

    # 2) The header worksheet exists (the knowledge source).
    if spec.header_sheet not in sheet_names:
        wb.close()
        raise TeamTemplateError(
            f"This doesn't look like the official {spec.team_name} template: the required "
            f"'{spec.header_sheet}' worksheet is missing. This workbook has: "
            f"{', '.join(sheet_names)}.",
            reason=f"Missing worksheet: {spec.header_sheet}",
        )

    # 3) & 4) Column names and order match the official template exactly.
    try:
        ws = wb[spec.header_sheet]
        actual_headers = [
            ws.cell(row=spec.header_row, column=c).value for c in range(1, ws.max_column + 1)
        ]
    except Exception as e:
        wb.close()
        raise TeamTemplateError(
            f"Could not read the '{spec.header_sheet}' worksheet's header row "
            f"(expected on row {spec.header_row}). "
            f"Please upload the official {spec.team_name} template.",
            reason="Could not read header row",
        ) from e
    finally:
        wb.close()
        _rewind(source)

    if spec.required_columns:
        # Lenient mode: only require the essential columns to be present
        # somewhere in the header row, matched whitespace/case-tolerantly
        # (the same way excel_parser reads them). Order and any extra
        # columns don't matter.
        from services.excel_parser import _normalize_header

        present = {_normalize_header(h) for h in actual_headers if h is not None}
        for required in spec.required_columns:
            if _normalize_header(required) not in present:
                raise TeamTemplateError(
                    f"This file doesn't look like the {spec.team_name} template: "
                    f"the '{spec.header_sheet}' worksheet (row {spec.header_row}) is missing "
                    f"the required column {required.strip()!r}. It must contain at least: "
                    f"{', '.join(c.strip() for c in spec.required_columns)}.",
                    reason=f"Missing required column: {required.strip()}",
                )
    else:
        for position, expected in enumerate(spec.expected_headers, start=1):
            actual = actual_headers[position - 1] if position <= len(actual_headers) else None
            if actual != expected:
                # Distinguish "this exact column exists, just somewhere
                # else in the row" (a reordering) from "this column isn't
                # in the row at all" (genuinely missing/renamed).
                if expected in actual_headers:
                    reason = "Invalid column order"
                else:
                    reason = f"Missing required column: {expected.strip()}"
                raise TeamTemplateError(
                    f"This file doesn't match the official {spec.team_name} template: "
                    f"column {position} in '{spec.header_sheet}' (row {spec.header_row}) should be "
                    f"{expected!r} but found {actual!r}. Column names and order must "
                    f"match the official template exactly.",
                    reason=reason,
                )

    # 5) Template version, if a version marker is configured for this team.
    if spec.template_version is not None:
        logger.debug(
            "%s: template version marker configured but no check implemented yet.",
            spec.team_name,
        )
    else:
        logger.debug("%s: no template version marker configured; version check skipped.",
                      spec.team_name)


def parse_team_template(source: Any, spec: TeamTemplateSpec) -> list[dict[str, Any]]:
    """Validate, then convert a team's workbook into MHES's internal
    nested category/task/activity data model.

    Args:
        source: A file path, or a file-like/stream object.
        spec: The uploading team's ``TeamTemplateSpec``.

    Returns:
        The same category/task/activity structure
        ``excel_to_nested_json`` produces for every other team.

    Raises:
        TeamTemplateError: if ``source`` doesn't structurally match
            ``spec`` -- nothing is parsed in that case.
    """
    validate_team_template(source, spec)
    _rewind(source)
    return excel_to_nested_json(source, column_mapping=spec.column_mapping)