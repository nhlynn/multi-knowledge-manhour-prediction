"""SSD Team's Excel-to-nested-JSON knowledge parser.

Independent from Bamawl/KiKan Team's config-driven import handling and
from SGL Team's own parser. SSD's official detail worksheet
(``詳細設計～システムテスト 本番移行`` in the SSD import/export template) has a
layout the generic single-header-row "phases mode" cannot express, and
one extra wrinkle no other team has:

1. **A THREE-row header.** Row 5 holds the field labels
   (``No.``/``機能名``/``機能概要``/``要件``/``難易度``/``新規・改定``/``見積根拠``)
   plus three merged group labels (``標準作業工数``, ``調整工数``,
   ``見積工数``). Row 6 holds sub-group labels (``ボリューム`` etc.). Row 7
   holds each group's own four phase sub-labels — ``詳細設計``/``実装``/
   ``単体テスト``/``結合テスト`` — repeated once under each of the three
   groups (so twelve phase-hour columns in all: J–M standard, N–Q
   adjustment, R–U estimate). SGL's own parser only has a TWO-row
   header; this one needs three.

2. **The standard-hours columns (J–M) are VLOOKUP formulas**, not
   literal values — ``=VLOOKUP($F<row>, 難易度別標準工数!$B$7:$G$10, ...)``
   keyed on each task's 難易度 (A/B/C/D/S). openpyxl reads a formula's
   *cached* value, which is lost the moment the workbook is re-saved by
   any tool that doesn't recalculate (openpyxl itself, this app's own
   sanitizer, etc.) — exactly the defect KiKan Team's shipped template
   suffers from. Rather than depend on that cache, this parser reads
   the ``難易度別標準工数`` lookup sheet DIRECTLY, builds an in-memory
   ``難易度 -> {phase: standard_hours}`` table, and resolves each task's
   standard hours itself. The adjustment columns (N–Q) are plain
   literal numbers and are read as-is. Final estimate per phase is then
   ``standard + adjustment``, computed here — never read from the R–U
   formula columns. This makes the import completely independent of any
   formula's cached value.

What IS reused (not reimplemented): once rows are folded into the same
``{category_slug: {"category": ..., "tasks": {task_key: {"task": ...,
"buffer_hours": ..., "activities": [...]}}}}`` accumulator shape
``services.excel_parser._process_phases_row`` builds, the exact same
``_build_nested_output``/``_log_conversion_summary`` helpers finish the
job — so the final nested JSON is byte-for-byte the same shape as every
other team's, and everything downstream (``EmbeddingService`` — text
extraction, embedding, FAISS indexing, metadata) needs no SSD-specific
awareness.

Per-task extra fields carried through generically (the same mechanism
SGL's ``work_detail`` uses — see ``services/search_service.py``'s
``_extra_task_fields``): ``difficulty`` (難易度), ``kind`` (新規・改定),
``block`` (which header block the task sits under), and each phase's
``adjustment_hours``. These survive search/Preview untouched and let
``services/ssd_export_builder.py`` write a re-exported task back with
its own 難易度 + adjustment so the template's own VLOOKUP + ``=J+N``
formulas recompute the standard and estimate natively on export.
"""

import logging
from typing import Any

import openpyxl

from services.excel_parser import (
    _build_nested_output,
    _log_conversion_summary,
    _safe_float,
    _slugify,
)

logger = logging.getLogger(__name__)

# The two worksheets this parser reads. Every OTHER sheet in the
# workbook (見積総額, 対応方針, 前提条件, システム構成図, シート説明, 体制,
# 成果物, スケジュール) is project-specific narrative/pricing/instruction,
# never man-hour knowledge, and is never touched here.
SSD_DETAIL_SHEET = "詳細設計～システムテスト 本番移行"
SSD_DETAIL_SHEET_2 = "詳細設計～システムテスト 本番移行_2"
SSD_LOOKUP_SHEET = "難易度別標準工数"

# --- 詳細 detail sheet fixed layout (1-indexed, matching openpyxl) ---
# Three-row header; task rows begin after it. These are structural
# facts about this one template (not user-configurable column_mapping),
# same as SGL's own _MAIN_HEADER_ROW/_PHASE_LABEL_ROW constants.
_FIELD_HEADER_ROW = 5      # No./機能名/機能概要/... + merged group labels
_PHASE_LABEL_ROW = 7       # 詳細設計/実装/単体テスト/結合テスト under each group
_DATA_START_ROW = 8        # first row that can hold a category header/task

# Field columns (row 5). Read by header text, not hardcoded index, so a
# reordered template is picked up — but the three-row structure itself
# is fixed knowledge, as above.
_FUNC_NAME_HEADER = "機能名"       # C — task name
_FUNC_OVERVIEW_HEADER = "機能概要"  # D — overview (embeddable)
_REQUIREMENT_HEADER = "要件（ユースケース）"  # E — requirement / use case (embeddable)
_DIFFICULTY_HEADER = "難易度"      # F — S/A/B/C/D, drives the lookup
_KIND_HEADER = "新規/改定"         # G — new vs revised
_BASIS_HEADER = "見積根拠"         # H — estimate basis (embeddable)

# The three merged group labels on row 5. The parser resolves all
# three phase-column groups from their merge spans:
#   標準作業工数 (J–M) — normally VLOOKUP formulas, but some rows enter a
#     literal standard value directly (e.g. 開発環境's 実装=2).
#   調整工数 (N–Q)     — literal manual +/- adjustments.
#   見積工数 (R–U)     — the TRUE final per-phase estimate. Normally
#     "=標準+調整", but special rows (機能外テスト/移行/管理 etc.) put a
#     direct value or formula here (e.g. コードインスペクション "=2/8",
#     本番化作業 literal 3). This column is the primary source of truth
#     for final hours; standard+adjustment is only a fallback for when
#     its cached value was lost to an openpyxl re-save.
_STANDARD_GROUP_HEADER = "標準作業工数（人日）"
_ADJUSTMENT_GROUP_HEADER = "調整工数（人日）"
_ESTIMATE_GROUP_HEADER = "見積工数（人日）"

# The four phase sub-labels, in template order, that appear under each
# group on row 7. Also the column order of the lookup sheet's own
# per-difficulty rows.
_PHASE_LABELS = ["詳細設計", "実装", "単体テスト", "結合テスト"]

# --- 難易度別標準工数 lookup sheet fixed layout ---
# Row 5 header: B=難易度 C=説明 D=詳細設計 E=実装 F=単体テスト G=結合テスト.
# Difficulty rows are 6–10 (S/A/B/C/D). "個別算出" (S) and blank (D)
# rows carry no numeric standard hours — treated as 0 here; such tasks
# still import if they have any nonzero adjustment.
_LOOKUP_DIFFICULTY_COL = 2   # B
_LOOKUP_FIRST_PHASE_COL = 4  # D (詳細設計); E/F/G follow in _PHASE_LABELS order
_LOOKUP_DATA_ROWS = range(6, 11)


def ssd_excel_to_nested_json(excel_path: str) -> list[dict[str, Any]]:
    """Convert SSD Team's detail worksheet(s) into the same nested JSON
    shape ``services.excel_parser.excel_to_nested_json`` produces for
    every other team.

    Reads ``難易度別標準工数`` first (as an in-memory lookup table), then
    both ``詳細設計～システムテスト 本番移行`` and its ``_2`` variant if present
    (the two proposal options ① / ②). A workbook missing the primary
    detail sheet returns an empty list rather than falling back to any
    other sheet.

    Row semantics:
      - A row in column C whose text starts with a section number
        (``1.``, ``2.``, ``3.`` …) is a CATEGORY header, forward-filled
        down to the tasks beneath it — not itself a task.
      - Any other row with non-blank 機能名 (C) is a real task.
      - Subtotal rows (``小計（工程別）``) and blank filler rows are
        skipped: they have no 機能名 text.
      - A task's standard hours per phase come from the lookup table
        (via its 難易度), NOT from the J–M formula columns. Its
        adjustment hours per phase come from the N–Q literal columns.
        Final per-phase estimate = standard + adjustment, computed here.
      - A task with every phase's final estimate at 0 is still imported
        as long as it has a 機能名 — SSD tasks legitimately net to 0 on
        a phase after a negative adjustment, and dropping them would
        lose real knowledge. (This differs from SGL, whose rows are
        gated on nonzero phase hours, because SSD's hours are derived,
        not literal.)
    """
    wb = openpyxl.load_workbook(excel_path, data_only=True)

    if SSD_DETAIL_SHEET not in wb.sheetnames:
        logger.warning(
            "SSD import: detail sheet %r not found in %s; nothing to import.",
            SSD_DETAIL_SHEET, excel_path,
        )
        return []

    if SSD_LOOKUP_SHEET not in wb.sheetnames:
        logger.warning(
            "SSD import: lookup sheet %r not found in %s; standard hours "
            "cannot be resolved. Nothing to import.",
            SSD_LOOKUP_SHEET, excel_path,
        )
        return []

    standard_lookup = _build_standard_lookup(wb[SSD_LOOKUP_SHEET])

    all_categories: dict[str, dict[str, Any]] = {}

    # Read both proposal sheets (① and ②) when present; tag each task
    # with its source block so identical function names across the two
    # approaches stay distinguishable downstream.
    sheet_names = [SSD_DETAIL_SHEET]
    if SSD_DETAIL_SHEET_2 in wb.sheetnames:
        sheet_names.append(SSD_DETAIL_SHEET_2)

    for sheet_name in sheet_names:
        _process_detail_sheet(
            wb[sheet_name], sheet_name, standard_lookup, all_categories,
        )

    result = _build_nested_output(all_categories)
    _log_conversion_summary(excel_path, result)
    return result


def _build_standard_lookup(ws) -> dict[str, dict[str, float]]:
    """Read the ``難易度別標準工数`` sheet into ``{difficulty: {phase:
    hours}}`` — the in-memory replacement for the detail sheet's
    VLOOKUP formulas.

    A difficulty row whose phase cell is non-numeric (S's "個別算出")
    or blank (D) yields 0.0 for that phase. Only literal cell values
    are read, so this never depends on any formula's cached result.
    """
    lookup: dict[str, dict[str, float]] = {}
    for row in _LOOKUP_DATA_ROWS:
        diff_val = ws.cell(row=row, column=_LOOKUP_DIFFICULTY_COL).value
        if diff_val is None or not str(diff_val).strip():
            continue
        difficulty = str(diff_val).strip()
        phases: dict[str, float] = {}
        for offset, label in enumerate(_PHASE_LABELS):
            col = _LOOKUP_FIRST_PHASE_COL + offset
            phases[label] = _safe_float(ws.cell(row=row, column=col).value)
        lookup[difficulty] = phases
    return lookup


def _resolve_field_columns(ws) -> dict[str, int]:
    """Map each field header on row 5 to its column index, by text —
    tolerant of a reordered template. Only the row-5 field labels are
    resolved here; the adjustment-group phase columns are resolved
    separately from the merge span (see ``_resolve_adjustment_columns``).
    """
    return {
        str(cell.value).strip(): cell.column
        for cell in ws[_FIELD_HEADER_ROW]
        if cell.value is not None and str(cell.value).strip()
    }


def _resolve_group_columns(
    ws, field_columns: dict[str, int], group_header: str,
) -> list[tuple[str, int]]:
    """Return ``[(phase_label, column_index), ...]`` for one merged
    phase-column group (標準作業工数 / 調整工数 / 見積工数), discovered from
    that group-label cell's merge span on row 5 and each column's own
    phase label on row 7 — never hardcoded, so a template that resizes
    or reorders a group is picked up automatically.
    """
    group_col = field_columns.get(group_header)
    if not group_col:
        return []

    end_col = group_col
    for merged in ws.merged_cells.ranges:
        if merged.min_row == _FIELD_HEADER_ROW and merged.min_col == group_col:
            end_col = merged.max_col
            break

    columns: list[tuple[str, int]] = []
    for col in range(group_col, end_col + 1):
        label = ws.cell(row=_PHASE_LABEL_ROW, column=col).value
        if label is not None and str(label).strip():
            columns.append((str(label).strip(), col))
    return columns


def _is_category_header(text: str) -> bool:
    """True if a 機能名-column cell is a section heading (``1.処理・画面``,
    ``3.全般`` …) rather than a real function name. These label the
    block their tasks sit under; forward-filled as the category.

    Accepts both ASCII and full-width digits (``１.`` as well as ``1.``)
    since SSD templates mix the two for section numbers.
    """
    stripped = text.strip()
    head = stripped.split(".", 1)[0]
    if not head:
        return False
    # str.isdigit() is True for full-width digits (１２３) too, so this
    # catches both "1.処理・画面" and "３.移行".
    return head.isdigit()


# The detail sheet repeats its whole field-header block (row 28 in the
# sample) above the second proposal's task list. A row whose 機能名 cell
# literally IS a field-header label is that repeated header, not a task
# — skip it. Matched against the row-5 field labels so a template that
# renames them stays covered.
_HEADER_LABELS = frozenset({
    _FUNC_NAME_HEADER, _FUNC_OVERVIEW_HEADER, _REQUIREMENT_HEADER,
    _DIFFICULTY_HEADER, _KIND_HEADER, _BASIS_HEADER, "No", "No.",
})


def _is_repeated_header(name: str, difficulty: str) -> bool:
    """True if a row is a repeated field-header block (e.g. the second
    proposal's own header at row 28) rather than a real task. The tell:
    its 機能名 cell holds a header label (``機能名``) and/or its 難易度
    cell holds the literal header text ``難易度``.
    """
    return name in _HEADER_LABELS or difficulty == _DIFFICULTY_HEADER


def _process_detail_sheet(
    ws,
    sheet_name: str,
    standard_lookup: dict[str, dict[str, float]],
    all_categories: dict[str, dict[str, Any]],
) -> None:
    """Fold every real task row on one detail sheet into
    ``all_categories`` (mutated in place, same accumulator shape
    ``services.excel_parser._process_phases_row`` builds)."""
    field_columns = _resolve_field_columns(ws)

    func_name_col = field_columns.get(_FUNC_NAME_HEADER)
    difficulty_col = field_columns.get(_DIFFICULTY_HEADER)
    if not (func_name_col and difficulty_col):
        logger.warning(
            "SSD import: expected headers (%r/%r) not found on row %d of "
            "sheet %r; skipping this sheet.",
            _FUNC_NAME_HEADER, _DIFFICULTY_HEADER, _FIELD_HEADER_ROW, sheet_name,
        )
        return

    overview_col = field_columns.get(_FUNC_OVERVIEW_HEADER)
    requirement_col = field_columns.get(_REQUIREMENT_HEADER)
    kind_col = field_columns.get(_KIND_HEADER)
    basis_col = field_columns.get(_BASIS_HEADER)
    standard_columns = _resolve_group_columns(ws, field_columns, _STANDARD_GROUP_HEADER)
    adjustment_columns = _resolve_group_columns(ws, field_columns, _ADJUSTMENT_GROUP_HEADER)
    estimate_columns = _resolve_group_columns(ws, field_columns, _ESTIMATE_GROUP_HEADER)

    current_category = ""

    def cell_text(row: int, col: int | None) -> str:
        if not col:
            return ""
        val = ws.cell(row=row, column=col).value
        return str(val).strip() if val is not None else ""

    def cell_number(row: int, col: int | None):
        """Return a cell's numeric value, or None if blank / lost-cache
        formula / non-numeric. Distinguishes a real 0 (present) from an
        absent value (None) — the discriminator the final-hours logic
        relies on."""
        if not col:
            return None
        val = ws.cell(row=row, column=col).value
        if val is None:
            return None
        try:
            return float(val)
        except (TypeError, ValueError):
            return None

    for row in range(_DATA_START_ROW, ws.max_row + 1):
        name = cell_text(row, func_name_col)
        if not name:
            continue  # subtotal / blank filler row

        if _is_category_header(name):
            current_category = name
            continue

        difficulty = cell_text(row, difficulty_col)

        if _is_repeated_header(name, difficulty):
            continue  # the second proposal's repeated field-header row

        if not current_category:
            continue  # a task before any category header — skip defensively

        std_phases = standard_lookup.get(difficulty, {})
        std_by_label = {label: col for label, col in standard_columns}
        adj_by_label = {label: col for label, col in adjustment_columns}
        est_by_label = {label: col for label, col in estimate_columns}

        # One record per phase carrying all three hour types the Excel
        # shows — 標準作業工数 / 調整工数 / 見積工数 — so search/Preview can
        # display the full breakdown, not just the final number.
        phase_records: list[dict[str, Any]] = []
        adjustments: dict[str, float] = {}

        # Iterate the phases in the ESTIMATE group's order (the four
        # 見積工数 sub-columns) — that's the authoritative per-phase final.
        phase_labels = [label for label, _ in estimate_columns] or _PHASE_LABELS

        for label in phase_labels:
            adjustment = cell_number(row, adj_by_label.get(label)) or 0.0
            adjustments[label] = adjustment

            # 標準作業工数: a literal in the J–M column (survives re-save,
            # e.g. 開発環境's 実装=2) wins; else the 難易度 lookup value.
            std_literal = cell_number(row, std_by_label.get(label))
            standard = std_literal if std_literal is not None else std_phases.get(label, 0.0)

            # 見積工数 (R–U) cached value is the file's own true final —
            # correct for BOTH normal rows ("=標準+調整") and special rows
            # that enter a direct value/formula here (機能外テスト/移行/管理).
            # A real 0 (cache present) is honored; only a genuinely absent
            # value (None: blank, or a formula whose cache was lost to an
            # openpyxl re-save) falls back to standard + adjustment.
            est_val = cell_number(row, est_by_label.get(label))
            estimate = est_val if est_val is not None else standard + adjustment

            phase_records.append({
                "label": label,
                "standard": standard,
                "adjustment": adjustment,
                "estimate": estimate,
            })

        _add_task(
            all_categories,
            category=current_category,
            task=name,
            phase_records=phase_records,
            adjustments=adjustments,
            difficulty=difficulty,
            kind=cell_text(row, kind_col),
            overview=cell_text(row, overview_col),
            requirement=cell_text(row, requirement_col),
            basis=cell_text(row, basis_col),
            block=sheet_name,
        )


def _build_embed_text(
    task: str, overview: str, requirement: str, basis: str,
) -> str:
    """Assemble the task's embeddable ``text`` from its descriptive
    columns (機能名 + 機能概要 + 要件 + 見積根拠), skipping any that are
    blank. This overrides the generic auto-generated summary text so
    semantic search matches on the function's actual description and
    domain vocabulary, not just an hours summary.
    """
    parts = [p for p in (task, overview, requirement, basis) if p]
    return " ".join(parts)


def _add_task(
    all_categories: dict[str, dict[str, Any]],
    *,
    category: str,
    task: str,
    phase_records: list[dict[str, Any]],
    adjustments: dict[str, float],
    difficulty: str,
    kind: str,
    overview: str,
    requirement: str,
    basis: str,
    block: str,
) -> None:
    """Fold one task row into ``all_categories``, creating the
    category/task entries as needed — same accumulator shape
    ``services.excel_parser._process_phases_row`` builds, plus SSD's
    own extra fields (difficulty/kind/block/per-phase adjustment) and
    an explicit embeddable ``text`` override.

    Each phase activity carries all three hour types the Excel shows —
    ``standard_hours`` (標準作業工数), ``adjustment_hours`` (調整工数), and
    ``estimate_hours`` (見積工数, the final) — so search/Preview can
    display the full breakdown. ``estimate_hours`` remains the value all
    totals sum over, so grand totals are unchanged.

    A task key already present (e.g. the same function name appearing on
    both proposal sheets) is disambiguated by block, so option ①'s and
    option ②'s versions of one function name stay separate rows rather
    than silently merging.
    """
    cat_slug = _slugify(category)
    task_slug = _slugify(task)

    if cat_slug not in all_categories:
        all_categories[cat_slug] = {"category": category, "tasks": {}}
    cat_data = all_categories[cat_slug]

    # Include block in the key so identical names across the two
    # proposal sheets don't collide into one task.
    block_slug = _slugify(block)
    task_key = f"{cat_slug}_{task_slug}_{block_slug}"

    if task_key not in cat_data["tasks"]:
        cat_data["tasks"][task_key] = {
            "task": task,
            "buffer_hours": 0.0,
            "activities": [],
            "difficulty": difficulty,
            "kind": kind,
            "block": block,
            "adjustment_hours": {},
            # 機能概要 (overview) and 要件 (requirement) are kept as their
            # own fields — in addition to being folded into the embeddable
            # `text` below — so the shared search service can match them
            # individually (see services/search_service.py's
            # _SEARCHABLE_TASK_TEXT_FIELDS), not only via semantic search.
            "overview": overview,
            "requirement": requirement,
            "text": _build_embed_text(task, overview, requirement, basis),
        }
    task_data = cat_data["tasks"][task_key]

    for rec in phase_records:
        label = rec["label"]
        activity_slug = _slugify(label)
        task_data["activities"].append({
            "id": f"{task_key}_{activity_slug}",
            "task_detail": label,
            # estimate_hours (見積工数) is the value totals sum over.
            "estimate_hours": rec["estimate"],
            # Extra per-phase breakdown fields — passed through generically
            # by services/excel_parser._build_activity_output and
            # services/search_service so the UI can show all three columns.
            "standard_hours": rec["standard"],
            "adjustment_hours": rec["adjustment"],
        })

    # Preserve per-phase adjustment at task level too, so export can
    # write 難易度 + 調整 back and let the template's own formulas
    # recompute standard + estimate.
    for label, adj in adjustments.items():
        task_data["adjustment_hours"].setdefault(label, adj)