"""KiKan Team's own Export Builder.

Independent from ``services/bamawl_export_builder.py`` (no shared code
between the two beyond the generic, team-agnostic helpers already in
``services/excel_parser.py``) -- built the same way Bamawl Team's
export was: directly on top of KiKan Team's single official Excel
workbook, ``import/kikan/kikan_import_template.xlsx`` -- the exact
same file Template Download serves and import validation accepts (see
``KikanExportBuilder.template_path`` below and
``utils/migrations/kikan_import_export_config.py``). There is
deliberately no separate import-only or export-only template file.

Design:

- ``工数詳細`` (Preview's Category -> Task -> Activity data, one
  selected function per row, using exactly the values/man-hours/
  remarks as edited in Preview -- nothing here re-derives or
  recomputes a number Preview already determined) and ``機能一覧`` are
  BOTH populated, kept in lockstep -- see the "機能一覧 sync" note
  below for why. ``Milestone`` and ``工数・費用`` ship exactly as the
  template has them.
- The exported workbook KEEPS the template's own auto-calculate
  formulas. On a populated ``工数詳細`` row, only ``実装工数``
  (Development, column F) is written as a literal -- the one phase the
  user edits in Preview. Every other phase column (``コードレビュー``,
  ``仕様理解``, ``QA``, ...) and the row's own ``合計(h)`` total is written
  as the template's OWN row-5 formula (``=F5*G$2``, chained ``=J5*K$2``,
  sum-based ``=(F5+H5+J5+L5)*O$2``, ``=SUM(F5:M5)*P$2``,
  ``=SUM(F5:P5)``), translated to that row (via
  ``openpyxl.formula.translate.Translator``), so opening the file and
  changing Development recomputes every derived phase and the total
  exactly as the pristine template does. This is safe because Preview
  already makes Development the only editable phase and derives every
  other phase from it via these same coefficients, so the formula
  always yields the number Preview showed. (Formula cells carry no
  cached value, so reading the file with ``data_only=True`` sees
  ``None`` until Excel opens and recalculates -- expected, and matches
  the original template.)
- A row this export does **not** populate is left FULLY blank: the
  clear step below blanks every phase column and ``合計(h)`` across the
  whole block up front, and formulas are re-written only into rows a
  selected function is actually written into -- so no phantom formula
  row is left for the workbook's built-in subtotal ranges (e.g.
  ``SUM(F5:F11)``) to pick up.
- ``業務分類`` (category) is merged across the whole function-row block
  in both worksheets (e.g. ``工数詳細``'s ``A5:A11``, ``機能一覧``'s
  ``A2:A8``) in the template -- only the merge's top-left cell is ever
  written (openpyxl requires this; the rest of a merged range must
  stay empty), so the merge itself is never touched/resized in either
  sheet. If the selected functions span more than one category, that
  single cell can't represent all of them -- the first one is used and
  a warning is logged (same "known limitation, documented rather than
  silently guessed around" approach ``bamawl_export_builder.py`` takes
  for its own edge cases).
- **機能一覧 sync**: the pristine template links the two worksheets --
  ``工数詳細``'s ``機能名称`` cell is originally a live
  ``=VLOOKUP(<this row's 機能ID>, 機能一覧!$D$2:$F$8, 2, FALSE)``
  formula, joining ``工数詳細``'s own function-ID column against
  ``機能一覧``'s ScreenID column to pull the function's real name from
  there. If only ``工数詳細`` were ever populated (leaving ``機能一覧``
  showing the template's own sample placeholder rows), that VLOOKUP
  would silently resolve to stale sample data -- or the two sheets
  would simply show two different, disconnected sets of functions with
  no relationship between them at all (a very confusing result for
  anyone reading ``機能一覧`` expecting it to describe what ``工数詳細``
  is estimating). So both sheets are written from the SAME
  ``tasks_with_category`` list, in the same row order, sharing one
  generated function-ID per row (``F001``, ``F002``, ...) written into
  both ``工数詳細``'s function-ID column and ``機能一覧``'s ScreenID
  column -- the exact join key the original VLOOKUP relationship used.
  ``工数詳細``'s ``機能名称`` cell itself is still written as a literal
  value (not restored as a live VLOOKUP formula) -- guaranteed correct
  regardless of any join-key mismatch, rather than depending on a
  formula recalculating correctly, while ``機能一覧`` having the
  matching real name at that same key means the *relationship* the
  template intends is still meaningfully true, just resolved once at
  export time instead of live in Excel.
  ``機能一覧``'s own ``機能ID`` and ``内容`` columns have no matching
  field in Preview's task data and are left blank, same reasoning
  ``Status`` is left blank in ``工数詳細`` below.
- A task's user-edited remarks have nowhere to go as a literal cell
  value -- ``工数詳細`` has no remarks/notes column of its own, and
  adding one would change the sheet's layout. Instead, remarks are
  attached as an Excel cell comment on the row's ``機能名称`` cell --
  carries the text without adding a visible column or altering the
  sheet's layout at all.
- ``Status`` (a dropdown-validated 大/中/小 field) is written from a
  task's own ``status`` field when present -- captured at import time
  from ``工数詳細``'s own ``Status`` column (see
  ``utils/migrations/kikan_import_export_config.py``'s
  ``extra_columns``) and carried through Preview/search generically,
  the same mechanism SGL's own ``work_detail``/``block`` fields use.
  Left blank only for a task with no such value (e.g. a brand-new
  function added directly in Preview, never imported from a
  workbook) -- same reasoning Bamawl's export blanks its
  ``ReqDefinition`` free-text sections.
- Unselected functions are never written: every row in the template's
  original function-row block (in EITHER worksheet) that isn't used by
  a selected function is cleared, not left with stale sample data.

**Known limitation** (a direct consequence of reusing this specific
template file rather than building a fresh one, same category of
limitation as Bamawl's own documented one): ``工数詳細``'s rollup rows
(person-hour/day/month sums, per-role breakdowns) are calibrated to the
template's own built-in 7-row function block, and ``機能一覧``'s own
block is expected to hold exactly as many rows. This module writes into
those existing blocks only (never shifting/extending either one), and
raises ``KikanExportError`` rather than overflow into ``工数詳細``'s
rollup rows if a project has more selected functions than the block
holds, or if ``機能一覧``'s own block turns out to have a different
number of rows than ``工数詳細``'s (the two must match for the
row-for-row sync above to make sense).
"""

import logging
import os
from typing import Any

import openpyxl
from openpyxl.comments import Comment
from openpyxl.formula.translate import Translator
from openpyxl.utils import get_column_letter

from services.base_export_service import BaseExportService, ExportContext
from services.excel_parser import _find_column, _normalize_header, _safe_float

logger = logging.getLogger(__name__)

_COMMENT_AUTHOR = "MHES"
_FUNCTION_LIST_SHEET = "機能一覧"
_FUNCTION_LIST_HEADER_ROW = 1


class KikanExportError(ValueError):
    """Raised when KiKan Team's export template can't be built from,
    or the system data doesn't fit it -- see ``build_kikan_workbook``."""


def _resolve_template_columns(ws, header_row: int) -> dict[str, int]:
    """Map each header cell's (stripped) name to its 1-indexed column
    position -- same approach ``bamawl_export_builder.py`` uses,
    reimplemented independently here rather than imported, so this
    module has no dependency on Bamawl's own export code."""
    name_to_col: dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        raw = ws.cell(row=header_row, column=c).value
        name_to_col[("" if raw is None else str(raw)).strip()] = c
    return name_to_col


def _column_index(name_to_col: dict[str, int], target: str | None) -> int | None:
    """Resolve a configured column name to its column index, tolerant
    of whitespace/case the same way the import side is."""
    if not target:
        return None
    matched = _find_column(list(name_to_col.keys()), target)
    return name_to_col.get(matched) if matched else None


def _template_capacity(ws, data_start_row: int, name_col: int, phase_cols: list[int]) -> int:
    """Return how many function rows are available below the header
    before the template's own rollup/summary block starts.

    Computed from the freshly-loaded, still-untouched template (before
    any clearing/writing happens below): a row with a real function
    name (here: a non-blank ``機能名称`` formula result) is counted; a
    row with that column blank is still counted *unless* one of the
    phase columns already has a value -- that's the signature of a
    rollup row (e.g. the person-hour/day/month sums below the real
    function rows), which marks the boundary.
    """
    row = data_start_row
    while row <= ws.max_row:
        name_val = ws.cell(row=row, column=name_col).value
        if not name_val:
            phase_has_value = any(
                ws.cell(row=row, column=c).value not in (None, "") for c in phase_cols
            )
            if phase_has_value:
                break
        row += 1
    return row - data_start_row


def _phase_value(activities: list[dict[str, Any]], label: str) -> float:
    """Return the estimate_hours of the activity matching ``label``
    (whitespace/case-insensitive), or 0.0 if this task has none."""
    target = _normalize_header(label)
    for act in activities:
        if _normalize_header(act.get("task_detail") or "") == target:
            return _safe_float(act.get("estimate_hours"))
    return 0.0


def _merged_block_row_span(ws, anchor_row: int, anchor_col: int) -> int | None:
    """Return the row span of the merged range anchored at
    ``(anchor_row, anchor_col)`` (e.g. ``機能一覧``'s ``A2:A8`` ->
    ``7``), or None if that cell isn't the top-left of any merge.
    """
    for merged_range in ws.merged_cells.ranges:
        if merged_range.min_row == anchor_row and merged_range.min_col == anchor_col:
            return merged_range.max_row - merged_range.min_row + 1
    return None


def _clear_block(ws, data_start_row: int, capacity: int, columns: list[int]) -> None:
    """Blank every cell in ``columns`` across the template's whole
    function-row block, before writing the selected functions into it
    -- so no leftover sample value (or stale comment) lingers past
    however many real rows are written below (mirrors
    ``bamawl_export_builder.py``'s own sample-row clearing step)."""
    for r in range(data_start_row, data_start_row + capacity):
        for c in columns:
            if c:
                cell = ws.cell(row=r, column=c)
                cell.value = None
                cell.comment = None


def build_kikan_workbook(
    filepath: str,
    categories: list[dict[str, Any]],
    column_mapping: dict[str, Any],
    template_path: str,
) -> None:
    """Populate KiKan Team's own Excel template's ``工数詳細`` AND
    ``機能一覧`` worksheets with Preview's data (kept in sync with each
    other -- see module docstring's "機能一覧 sync" note) and save the
    result to ``filepath``. Neither ``Milestone`` nor ``工数・費用`` is
    touched.

    Args:
        filepath: Where to save the populated workbook.
        categories: The Preview page's Category -> Task -> Activity
            data (same shape ``services/export_workbook_service.py``
            receives) -- the single source of truth for what's
            written. Every task across every category is one selected
            function; only these are exported, one per ``工数詳細``
            row, in order.
        column_mapping: KiKan Team's configured phases-mode column
            mapping (``sheet``, ``header_row``, ``task_column``,
            ``category_column``, ``id_column``, ``phase_columns`` --
            see ``utils/migrations/kikan_import_export_config.py``).
        template_path: Path to KiKan Team's single official template
            workbook (``import/kikan/kikan_import_template.xlsx``).

    Raises:
        KikanExportError: if the template file/worksheet/columns can't
            be found or understood, or the project has more selected
            functions than the template's function-row block can hold
            without touching its rollup rows (see module docstring).
    """
    if not os.path.isfile(template_path):
        raise KikanExportError(f"KiKan Team's export template file is missing: {template_path}")

    sheet_name = column_mapping.get("sheet")
    header_row = column_mapping.get("header_row") or 1

    wb = openpyxl.load_workbook(template_path)

    if sheet_name not in wb.sheetnames:
        raise KikanExportError(
            f"KiKan Team's export template is missing the required '{sheet_name}' worksheet."
        )
    ws = wb[sheet_name]

    name_to_col = _resolve_template_columns(ws, header_row)
    name_col = _column_index(name_to_col, column_mapping.get("task_column"))
    if name_col is None:
        raise KikanExportError(
            "KiKan Team's export template's function-name column could not be located; "
            "cannot populate 工数詳細."
        )
    category_col = _column_index(name_to_col, column_mapping.get("category_column"))
    no_col = _column_index(name_to_col, column_mapping.get("id_column"))
    func_id_col = _column_index(name_to_col, "機能ID")
    status_col = _column_index(name_to_col, "Status")

    phase_cols = [
        (phase["label"], _column_index(name_to_col, phase["column"]))
        for phase in column_mapping.get("phase_columns", [])
    ]
    phase_cols = [(label, idx) for label, idx in phase_cols if idx is not None]
    if not phase_cols:
        raise KikanExportError(
            "KiKan Team's export template's phase-hour columns could not be located; "
            "cannot populate 工数詳細."
        )
    # The exported workbook now KEEPS the template's own auto-calculate
    # formulas: of all the phase columns, only "Development" (実装工数,
    # column F) is a literal the user edited in Preview; every other
    # phase column (コードレビュー, 仕様理解, QA, ...) and the 合計(h) total
    # is one of the template's own row-5 formulas (e.g. G5=F5*G$2,
    # chained J5*K$2, sum-based (F5+H5+J5+L5)*O$2, SUM(F5:M5)*P$2,
    # Q5=SUM(F5:P5)). Each is captured below and re-written (translated)
    # into every populated row, so changing Development in Excel
    # recomputes the rest exactly as the pristine template does. Because
    # Preview already derives every non-Development phase from Development
    # via these same coefficients, the formula always yields the number
    # Preview showed.
    dev_col = next(
        (idx for label, idx in phase_cols if _normalize_header(label) == _normalize_header("Development")),
        None,
    )
    total_col = _column_index(name_to_col, column_mapping.get("total_column"))

    tasks_with_category = [
        (cat.get("category", ""), task) for cat in categories for task in cat.get("tasks", [])
    ]

    data_start_row = header_row + 1
    phase_col_indexes = [idx for _label, idx in phase_cols]
    capacity = _template_capacity(ws, data_start_row, name_col, phase_col_indexes)
    if len(tasks_with_category) > capacity:
        raise KikanExportError(
            f"This project has {len(tasks_with_category)} function(s), but KiKan Team's "
            f"export template's '{sheet_name}' worksheet only has room for {capacity} "
            f"before its built-in rollup rows -- reduce the number of selected functions, "
            f"or update the template."
        )

    # Capture each derived phase column's and the 合計(h) total column's
    # template formula BEFORE the block is cleared below, so each can be
    # re-written (row-translated) into every populated row. Only genuine
    # formulas (leading "=") are captured; a column that isn't a formula
    # there falls back to the old literal behavior.
    #
    # Source row: capture from the second data row when the block has one
    # (KiKan's rows are uniform -- F6=``=F6*G$2`` -- but this also makes
    # the two builders share one rule; the sibling Bamawl builder MUST
    # skip its first data row, whose coefficient refs are relative). Any
    # subsequent row translates correctly to any other row (including
    # back up to the first).
    formula_src_row = data_start_row
    if capacity >= 2 and total_col:
        probe = ws.cell(row=data_start_row + 1, column=total_col).value
        if isinstance(probe, str) and probe.startswith("="):
            formula_src_row = data_start_row + 1

    captured_formulas: dict[int, str] = {}
    for col_idx in [idx for _label, idx in phase_cols if idx != dev_col] + (
        [total_col] if total_col else []
    ):
        raw = ws.cell(row=formula_src_row, column=col_idx).value
        if isinstance(raw, str) and raw.startswith("="):
            captured_formulas[col_idx] = raw

    def _row_formula(col_idx: int, r: int) -> str:
        """Translate a captured template formula from ``formula_src_row``
        to row ``r`` (absolute refs like ``G$2`` stay fixed; relative
        refs like ``F5`` shift to ``F{r}``)."""
        origin = f"{get_column_letter(col_idx)}{formula_src_row}"
        dest = f"{get_column_letter(col_idx)}{r}"
        return Translator(captured_formulas[col_idx], origin=origin).translate_formula(dest)

    # 機能一覧 sync (see module docstring's "機能一覧 sync" note) -- both
    # sheets are populated together from the same tasks_with_category
    # list, so the pristine template's 工数詳細 <-> 機能一覧 relationship
    # (originally a live VLOOKUP) stays meaningfully true after export
    # instead of leaving 機能一覧 orphaned on stale sample data.
    if _FUNCTION_LIST_SHEET not in wb.sheetnames:
        raise KikanExportError(
            f"KiKan Team's export template is missing the required "
            f"'{_FUNCTION_LIST_SHEET}' worksheet."
        )
    func_list_ws = wb[_FUNCTION_LIST_SHEET]
    func_list_cols = _resolve_template_columns(func_list_ws, _FUNCTION_LIST_HEADER_ROW)
    func_list_data_start_row = _FUNCTION_LIST_HEADER_ROW + 1
    func_list_category_col = _column_index(func_list_cols, "業務分類")
    func_list_no_col = _column_index(func_list_cols, "番号")
    func_list_id_col = _column_index(func_list_cols, "機能ID")
    func_list_screen_id_col = _column_index(func_list_cols, "ScreenID")
    func_list_name_col = _column_index(func_list_cols, "機能名称")
    func_list_content_col = _column_index(func_list_cols, "内容")
    if func_list_name_col is None or func_list_screen_id_col is None:
        raise KikanExportError(
            f"KiKan Team's export template's '{_FUNCTION_LIST_SHEET}' worksheet is missing "
            f"its 機能名称/ScreenID columns; cannot keep it in sync with '{sheet_name}'."
        )

    func_list_capacity = _merged_block_row_span(
        func_list_ws, func_list_data_start_row, func_list_category_col,
    ) if func_list_category_col else None
    if func_list_capacity is not None and func_list_capacity != capacity:
        raise KikanExportError(
            f"KiKan Team's export template's '{_FUNCTION_LIST_SHEET}' worksheet has room for "
            f"{func_list_capacity} function row(s), but '{sheet_name}' has room for "
            f"{capacity} -- the two must match for exported functions to stay in sync "
            f"across both worksheets. Update the template so both blocks are the same size."
        )

    # Clear the template's whole function-row block first -- 機能名称
    # (whose original VLOOKUP formula is replaced with a literal name
    # only for rows a selected function is actually written into),
    # 番号/機能ID/Status, and EVERY phase column plus 合計(h) -- so no
    # leftover sample name/hours, and no template formula, lingers past
    # however many real rows are written below. A row this export
    # doesn't populate is left fully blank (no phantom formula), so the
    # workbook's built-in subtotal ranges (e.g. SUM(F5:F11)) don't pick
    # up empty formula rows; formulas are re-written (translated) only
    # into rows a selected function is actually written into.
    #
    # category_col is excluded here -- it's the top-left cell of a
    # merge spanning the whole block (A5:A11); every other cell in that
    # merged range is a read-only MergedCell placeholder, not a real
    # cell, and is cleared/set only once below, at the merge's top-left
    # row.
    _clear_block(
        ws, data_start_row, capacity,
        [name_col, no_col, func_id_col, status_col, total_col, *phase_col_indexes],
    )
    if category_col:
        ws.cell(row=data_start_row, column=category_col).value = None

    # Same clearing for 機能一覧's own block, keyed by its own column
    # set (機能一覧 has no phase/hours columns, so nothing there ever
    # needs the "blank one ratio-formula base column" treatment).
    _clear_block(
        func_list_ws, func_list_data_start_row, capacity,
        [func_list_no_col, func_list_id_col, func_list_screen_id_col,
         func_list_name_col, func_list_content_col],
    )
    if func_list_category_col:
        func_list_ws.cell(row=func_list_data_start_row, column=func_list_category_col).value = None

    categories_used = {cat for cat, _task in tasks_with_category if cat}
    if len(categories_used) > 1:
        logger.warning(
            "KiKan export: selected functions span %d categories (%s), but '%s' can only "
            "show one category label for the whole block (merged cell) -- using %r.",
            len(categories_used), sorted(categories_used), sheet_name,
            next(iter(categories_used)),
        )
    block_category = next((cat for cat, _task in tasks_with_category if cat), None)
    if category_col and block_category:
        ws.cell(row=data_start_row, column=category_col, value=block_category)
    if func_list_category_col and block_category:
        func_list_ws.cell(row=func_list_data_start_row, column=func_list_category_col, value=block_category)

    unmatched_labels: set[str] = set()
    configured_labels = {_normalize_header(label) for label, _idx in phase_cols}

    for i, (_category, task) in enumerate(tasks_with_category, start=1):
        row = data_start_row + i - 1
        func_list_row = func_list_data_start_row + i - 1
        activities = task.get("activities", []) or []
        task_name = task.get("task", "")

        for act in activities:
            norm = _normalize_header(act.get("task_detail") or "")
            if norm and norm not in configured_labels:
                unmatched_labels.add(act.get("task_detail"))

        # Shared join key written into BOTH 工数詳細's function-ID
        # column and 機能一覧's ScreenID column -- see the module
        # docstring's "機能一覧 sync" note for why: this is the exact
        # key the pristine template's own VLOOKUP joined the two
        # sheets on. A task that came from an actual KiKan import
        # already carries its OWN real ``screen_id`` (see
        # services/kikan_import_parser.py) -- reused as-is so a
        # round-tripped function keeps the same identity across
        # import -> Preview -> export. Only a task with no such value
        # (e.g. a brand-new function added directly in Preview, never
        # imported from a workbook) falls back to a freshly generated
        # placeholder.
        shared_func_id = task.get("screen_id") or f"F{i:03d}"

        if no_col:
            ws.cell(row=row, column=no_col, value=i)
        if func_id_col:
            ws.cell(row=row, column=func_id_col, value=shared_func_id)
        if status_col:
            status = task.get("status")
            if status:
                ws.cell(row=row, column=status_col, value=status)

        name_cell = ws.cell(row=row, column=name_col, value=task_name)
        remarks = task.get("remarks")
        if remarks:
            name_cell.comment = Comment(str(remarks), _COMMENT_AUTHOR)

        if func_list_no_col:
            func_list_ws.cell(row=func_list_row, column=func_list_no_col, value=i)
        func_list_ws.cell(row=func_list_row, column=func_list_screen_id_col, value=shared_func_id)
        func_list_ws.cell(row=func_list_row, column=func_list_name_col, value=task_name)
        # 機能一覧's OWN 機能ID/内容 -- different columns from anything
        # on 工数詳細, only ever available for a task that came from an
        # actual KiKan import (see services/kikan_import_parser.py);
        # left blank for a brand-new Preview-only task, same reasoning
        # Status is left blank below when absent.
        function_id = task.get("function_id")
        if func_list_id_col and function_id:
            func_list_ws.cell(row=func_list_row, column=func_list_id_col, value=function_id)
        content = task.get("content")
        if func_list_content_col and content:
            func_list_ws.cell(row=func_list_row, column=func_list_content_col, value=content)

        # Development (実装工数, the base) stays a literal -- the one
        # phase the user edited in Preview -- written even when 0 (blank
        # for 0 is fine; the derived formulas then compute 0). Every
        # other phase column, and 合計(h), is written as the template's
        # own formula translated to this row, so the exported file
        # recomputes them live in Excel (changing Development
        # recalculates the rest) exactly as the pristine template does.
        # Because Preview derives every non-Development phase from
        # Development via these same coefficients, the formula always
        # yields the number Preview showed. A derived/total column with
        # no captured template formula falls back to a literal.
        #
        # Direct attribute assignment (not the value= kwarg) because
        # ws.cell(row, column, value=None) leaves a cell untouched when
        # value is None -- here every cell was already blanked by the
        # clear step, so a literal-0 phase correctly stays blank.
        for label, col_idx in phase_cols:
            if col_idx == dev_col or col_idx not in captured_formulas:
                ws.cell(row=row, column=col_idx).value = _phase_value(activities, label) or None
            else:
                ws.cell(row=row, column=col_idx).value = _row_formula(col_idx, row)

        if total_col is not None:
            if total_col in captured_formulas:
                ws.cell(row=row, column=total_col).value = _row_formula(total_col, row)
            else:
                ws.cell(row=row, column=total_col).value = (
                    sum(_phase_value(activities, lbl) for lbl, _ in phase_cols) or None
                )

    if unmatched_labels:
        logger.warning(
            "KiKan export: %d activity label(s) didn't match any configured phase column "
            "and were left out of '%s': %s",
            len(unmatched_labels), sheet_name, sorted(unmatched_labels),
        )

    wb.save(filepath)
    logger.info(
        "Built KiKan Team export workbook: %s (%d selected function(s) written into '%s' "
        "and kept in sync with '%s').",
        filepath, len(tasks_with_category), sheet_name, _FUNCTION_LIST_SHEET,
    )


class KikanExportBuilder(BaseExportService):
    """KiKan Team's export builder (Strategy Pattern) -- the single
    home for everything KiKan-specific about exporting: the Strategy
    Pattern wiring (``build``), and how to resolve KiKan Team's own
    ``column_mapping``/template path (``resolve_column_mapping``,
    ``template_path``), which used to live in ``routes/export.py`` as
    KiKan-only helper functions -- mirroring exactly how
    ``BamawlExportBuilder`` (``services/bamawl_export_builder.py``)
    consolidates the same shape of logic for Bamawl Team.

    Independent from ``BamawlExportBuilder`` -- no shared code between
    the two beyond the generic, team-agnostic ``BaseExportService`` and
    ``services/excel_parser.py`` helpers both already used. Nothing
    here imports from, or is imported by,
    ``services/bamawl_export_builder.py``.

    ``build`` itself still simply delegates to ``build_kikan_workbook``
    above (unchanged) -- this class is a thin, dedicated container
    around KiKan's own already-existing, already-tested logic, not a
    reimplementation of it.
    """

    team_name = "KiKan Team"

    @staticmethod
    def resolve_column_mapping(mhes_db_path: str, team_id: int) -> dict[str, Any] | None:
        """Return KiKan Team's configured import column mapping for
        ``team_id``, or None if it hasn't been seeded yet.

        The export builder reuses this (rather than a separate
        config) — it already describes exactly which worksheet/columns
        ``工数詳細``'s data lives in, the same mapping
        ``services/team_template_validator.py`` reads it with.
        """
        from repositories.team_import_config_repository import TeamImportConfigRepository

        repo = TeamImportConfigRepository(mhes_db_path)
        config = repo.get_by_team_id(team_id)
        return config["column_mapping"] if config else None

    @staticmethod
    def template_path(app_root_path: str) -> str:
        """Path to KiKan Team's single official Excel template.

        ``import/kikan/kikan_import_template.xlsx`` -- the same public
        sample workbook downloaded from Template Download and accepted
        by import validation -- is the *only* official KiKan Team
        template: there is deliberately no separate, internal-only
        export template. Export builds directly on top of this exact
        file (copy, then populate, then save), the same "one workbook
        for everything" design Bamawl Team's own template uses.

        ``simple_resource/kikan_import_export_template.xlsx`` (the
        original real-data workbook this public template was generated
        from -- see ``import/kikan/build_sample_template.py``) is no
        longer read anywhere at runtime; it exists on disk only as that
        script's source input, kept for reproducibility if the sample
        ever needs regenerating.
        """
        return os.path.join(app_root_path, "import", "kikan", "kikan_import_template.xlsx")

    def build(self, context: ExportContext) -> None:
        build_kikan_workbook(
            context.filepath, context.categories, context.column_mapping, context.template_path,
        )