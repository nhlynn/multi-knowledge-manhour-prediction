"""SGL Team's Excel-to-nested-JSON knowledge parser.

Independent from Bamawl/KiKan Team's import handling: those two teams
are entirely config-driven through ``services/excel_parser.py``'s
generic "phases mode" (a single header row, column names resolved by
text match). SGL's official worksheet (``詳細見積_マスタと予実比較`` in
``simple_resource/sgl_import_export_format.xlsx``) has a genuinely
different shape that mode can't express: a header split across TWO
rows (row 2 holds "No."/"区分"/"項目"/... and the merged "工数（人時間）"
group label; row 3 holds each phase's own sub-label -- "要件定義",
"設計", "開発", "テスト", "クラウド対応", "その他" -- underneath it).
Rather than bend the shared single-header-row pipeline to fit one
team's layout, this module reads the sheet directly.

What IS reused (not reimplemented): once rows are folded into the same
``{category_slug: {"category": ..., "tasks": {task_key: {"task": ...,
"buffer_hours": ..., "activities": [...]}}}}`` accumulator shape
``services.excel_parser._process_phases_row`` builds, the exact same
``_build_nested_output``/``_log_conversion_summary`` helpers finish the
job -- so the final nested JSON (categories/tasks/activities, their
embedding-ready ``text`` fields, summary math) is byte-for-byte the
same shape as Bamawl/KiKan/every other team's, and everything
downstream (``EmbeddingService.process_excel_file`` — text extraction,
embedding generation, FAISS indexing, metadata) needs no SGL-specific
awareness at all.
"""

import logging
from typing import Any

import openpyxl

from services.excel_parser import _build_nested_output, _log_conversion_summary, _safe_float, _slugify

logger = logging.getLogger(__name__)

# The ONLY worksheet ever read for SGL knowledge import -- "見積・金額サマリ"
# (the other sheet in the workbook, a summary/amount rollup) is never
# touched here.
SGL_SHEET_NAME = "詳細見積_マスタと予実比較"

# Fixed knowledge about this one template's layout (not user-configurable
# column_mapping data, since the two-row header can't be expressed that
# way) -- row numbers are 1-indexed, matching openpyxl.
_MAIN_HEADER_ROW = 2
_PHASE_LABEL_ROW = 3
_DATA_START_ROW = 4

_CATEGORY_HEADER = "区分"
_TASK_HEADER = "項目"
_PHASE_GROUP_HEADER = "工数（人時間）"
_WORK_DETAIL_HEADER = "作業詳細"


def sgl_excel_to_nested_json(excel_path: str) -> list[dict[str, Any]]:
    """Convert SGL Team's detail worksheet into the same nested JSON
    shape ``services.excel_parser.excel_to_nested_json`` produces for
    every other team.

    Reads ONLY ``SGL_SHEET_NAME`` -- if the workbook doesn't contain it,
    returns an empty list rather than falling back to any other sheet.

    Row semantics (mirroring ``_process_phases_row``'s own rules, so a
    hand-filled sample workbook behaves the same way regardless of
    which parser reads it):
      - "区分" (category) is forward-filled down blank rows -- filled in
        only once per group.
      - A row counts as a real task only if "項目" is non-blank AND at
        least one phase sub-column is greater than 0 -- both group
        subtotal/rollup rows (always blank "項目") and blank filler
        rows (blank "項目" and/or all-zero phase hours) are skipped,
        not imported as tasks. Deliberately does NOT gate on "No."
        (which holds a live ``=ROW()-4`` formula): a workbook re-saved
        by any tool that doesn't recalculate formulas (e.g. this app's
        own ``import/sgl/build_sample_template.py``, or openpyxl in
        general) loses every formula's *cached* value workbook-wide,
        which would otherwise make every row look like it has no
        "No." at all and silently import zero tasks -- the two checks
        above are sufficient on their own and depend only on literal
        cell values, never a formula's cached result.
      - Each of the six phase sub-columns (要件定義/設計/開発/テスト/
        クラウド対応/その他) with a value greater than 0 becomes its own
        Activity Detail under that task -- a row with no phase hours at
        all is skipped entirely.

    Args:
        excel_path: Path to the SGL Excel file.

    Returns:
        List of category-level dictionaries, one per category --
        identical shape to every other team's nested JSON.
    """
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    if SGL_SHEET_NAME not in wb.sheetnames:
        logger.warning(
            "SGL import: sheet %r not found in %s; nothing to import.",
            SGL_SHEET_NAME, excel_path,
        )
        return []

    ws = wb[SGL_SHEET_NAME]

    header_columns = {
        str(cell.value).strip(): cell.column
        for cell in ws[_MAIN_HEADER_ROW]
        if cell.value is not None
    }
    category_col = header_columns.get(_CATEGORY_HEADER)
    task_col = header_columns.get(_TASK_HEADER)
    phase_group_col = header_columns.get(_PHASE_GROUP_HEADER)
    work_detail_col = header_columns.get(_WORK_DETAIL_HEADER)

    if not (category_col and task_col and phase_group_col):
        logger.warning(
            "SGL import: expected headers (%r/%r/%r) not all found on "
            "row %d of sheet %r in %s; nothing to import.",
            _CATEGORY_HEADER, _TASK_HEADER, _PHASE_GROUP_HEADER,
            _MAIN_HEADER_ROW, SGL_SHEET_NAME, excel_path,
        )
        return []

    phase_columns = _resolve_phase_columns(ws, phase_group_col)
    if not phase_columns:
        logger.warning(
            "SGL import: no phase sub-columns found under %r on row %d "
            "of sheet %r in %s; nothing to import.",
            _PHASE_GROUP_HEADER, _PHASE_LABEL_ROW, SGL_SHEET_NAME, excel_path,
        )
        return []

    all_categories: dict[str, dict[str, Any]] = {}
    current_category = ""

    for row in range(_DATA_START_ROW, ws.max_row + 1):
        category_val = ws.cell(row=row, column=category_col).value
        if category_val is not None and str(category_val).strip():
            current_category = str(category_val).strip()

        task_val = ws.cell(row=row, column=task_col).value
        task = str(task_val).strip() if task_val is not None else ""
        if not task or not current_category:
            continue

        row_activities = [
            (label, _safe_float(ws.cell(row=row, column=col).value))
            for label, col in phase_columns
        ]
        row_activities = [(label, hours) for label, hours in row_activities if hours > 0]
        if not row_activities:
            continue

        row_work_detail = ""
        if work_detail_col:
            wd_val = ws.cell(row=row, column=work_detail_col).value
            row_work_detail = str(wd_val).strip() if wd_val is not None else ""

        _add_task_activities(all_categories, current_category, task, row_activities, row_work_detail)

    result = _build_nested_output(all_categories)
    _log_conversion_summary(excel_path, result)
    return result


def _resolve_phase_columns(ws, phase_group_col: int) -> list[tuple[str, int]]:
    """Return ``[(phase_label, column_index), ...]`` for the phase
    sub-columns merged under the "工数（人時間）" header cell.

    Reads the merge span of that header cell (row ``_MAIN_HEADER_ROW``)
    to find how many columns it spans, then reads each of those
    columns' own label from ``_PHASE_LABEL_ROW`` -- so a future edit
    reordering or resizing the phase columns is picked up automatically
    rather than assuming a fixed count/position.
    """
    end_col = phase_group_col
    for merged_range in ws.merged_cells.ranges:
        if merged_range.min_row == _MAIN_HEADER_ROW and merged_range.min_col == phase_group_col:
            end_col = merged_range.max_col
            break

    phase_columns = []
    for col in range(phase_group_col, end_col + 1):
        label = ws.cell(row=_PHASE_LABEL_ROW, column=col).value
        if label is not None and str(label).strip():
            phase_columns.append((str(label).strip(), col))
    return phase_columns


def _add_task_activities(
    all_categories: dict[str, dict[str, Any]],
    category: str,
    task: str,
    row_activities: list[tuple[str, float]],
    row_work_detail: str = "",
) -> None:
    """Fold one task row's phase activities (and 作業詳細 text, if any)
    into ``all_categories``, creating the category/task entries as
    needed -- same accumulator shape
    ``services.excel_parser._process_phases_row`` builds, plus SGL's
    own ``work_detail`` field.

    A task spanning multiple rows accumulates ``row_work_detail`` from
    each row it appears on, joined by newline, in row order -- mirrors
    how a single task's phase-hour activities themselves already
    accumulate across its rows.
    """
    cat_slug = _slugify(category)
    task_slug = _slugify(task)

    if cat_slug not in all_categories:
        all_categories[cat_slug] = {"category": category, "tasks": {}}
    cat_data = all_categories[cat_slug]

    task_key = f"{cat_slug}_{task_slug}"
    if task_key not in cat_data["tasks"]:
        cat_data["tasks"][task_key] = {
            "task": task, "buffer_hours": 0.0, "activities": [], "work_detail": "",
        }
    task_data = cat_data["tasks"][task_key]

    for label, hours in row_activities:
        activity_slug = _slugify(label)
        task_data["activities"].append({
            "id": f"{cat_slug}_{task_slug}_{activity_slug}",
            "task_detail": label,
            "estimate_hours": hours,
        })

    if row_work_detail:
        task_data["work_detail"] = (
            f"{task_data['work_detail']}\n{row_work_detail}"
            if task_data["work_detail"] else row_work_detail
        )